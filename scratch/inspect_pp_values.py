import openpyxl

wb = openpyxl.load_workbook("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=True)

# Let's inspect 3.Normalisasi
ws_norm = wb["3.Normalisasi"]
# A06 is alternative 6, which is row 10 (since row 5 is A01, row 10 is A06)
# K12 is Approx price (column 13)
norm_a06_k12 = ws_norm.cell(row=10, column=13).value
# A01 is row 5, K2 is Weight oz (column 3)
norm_a01_k2 = ws_norm.cell(row=5, column=3).value

# Let's inspect 5.Pref.Nyata
ws_pp = wb["5.Pref.Nyata"]
pp_a06_k12 = ws_pp.cell(row=10, column=13).value
pp_a01_k2 = ws_pp.cell(row=5, column=3).value

# Let's inspect 4.Pref.Teoritis
ws_tp = wb["4.Pref.Teoritis"]
tp_k12 = ws_tp.cell(row=5, column=13).value # constant across rows
tp_k2 = ws_tp.cell(row=5, column=3).value

print(f"Norm A06 K12: {norm_a06_k12}")
print(f"Norm A01 K2: {norm_a01_k2}")
print(f"TP K12: {tp_k12}")
print(f"TP K2: {tp_k2}")
print(f"PP A06 K12: {pp_a06_k12}")
print(f"PP A01 K2: {pp_a01_k2}")
