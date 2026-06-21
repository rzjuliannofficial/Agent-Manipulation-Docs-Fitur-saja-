import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

def copy_cell_style(src_cell, dest_cell):
    if src_cell.font:
        dest_cell.font = copy.copy(src_cell.font)
    if src_cell.border:
        dest_cell.border = copy.copy(src_cell.border)
    if src_cell.fill:
        dest_cell.fill = copy.copy(src_cell.fill)
    if src_cell.alignment:
        dest_cell.alignment = copy.copy(src_cell.alignment)
    dest_cell.number_format = src_cell.number_format

def copy_row_style(ws, src_row, dest_row, num_cols):
    for col in range(1, num_cols + 1):
        copy_cell_style(ws.cell(row=src_row, column=col), ws.cell(row=dest_row, column=col))

def fix_references():
    wb = openpyxl.load_workbook("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
    
    # 1. Update '3.Normalisasi'
    ws_norm = wb['3.Normalisasi']
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        ws_norm.cell(row=3, column=col, value=f"='1.Data'!{col_letter}30")
        ws_norm.cell(row=4, column=col, value=f"='1.Data'!{col_letter}31")
        
    for idx in range(20):
        row_num = 6 + idx
        alt_code = f"A{idx+1:02d}"
        ws_norm.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            data_row = 9 + idx
            if col in [2, 3, 4, 5]: # Benefit
                ws_norm.cell(row=row_num, column=col, value=f"=IF(${col_letter}$3=${col_letter}$4,0,('1.Data'!{col_letter}{data_row}-${col_letter}$4)/(${col_letter}$3-${col_letter}$4))")
            else: # Cost
                ws_norm.cell(row=row_num, column=col, value=f"=IF(${col_letter}$3=${col_letter}$4,0,(${col_letter}$3-'1.Data'!{col_letter}{data_row})/(${col_letter}$3-${col_letter}$4))")

    # 2. Update '4.Pref.Teoritis'
    ws_pref_t = wb['4.Pref.Teoritis']
    for idx in range(20):
        row_num = 5 + idx
        alt_code = f"A{idx+1:02d}"
        ws_pref_t.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            ws_pref_t.cell(row=row_num, column=col, value=f"='2.SWARA'!F{col+17}/20")

    # 3. Update '5.Pref.Nyata'
    ws_pref_n = wb['5.Pref.Nyata']
    for idx in range(20):
        row_num = 4 + idx
        alt_code = f"A{idx+1:02d}"
        ws_pref_n.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            pref_t_row = 5 + idx
            norm_row = 6 + idx
            ws_pref_n.cell(row=row_num, column=col, value=f"='4.Pref.Teoritis'!{col_letter}{pref_t_row}*'3.Normalisasi'!{col_letter}{norm_row}")

    # 4. Update '6.Gap'
    ws_gap = wb['6.Gap']
    for idx in range(20):
        row_num = 4 + idx
        alt_code = f"A{idx+1:02d}"
        ws_gap.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            pref_t_row = 5 + idx
            pref_n_row = 4 + idx
            ws_gap.cell(row=row_num, column=col, value=f"='4.Pref.Teoritis'!{col_letter}{pref_t_row}-'5.Pref.Nyata'!{col_letter}{pref_n_row}")
        ws_gap.cell(row=row_num, column=12, value=f"=SUM(B{row_num}:K{row_num})")
        ws_gap.cell(row=row_num, column=13, value=f"=RANK(L{row_num},$L$4:$L$23,1)")

    # 5. Update '7.Hasil'
    ws_hasil = wb['7.Hasil']
    for idx in range(20):
        row_num = 4 + idx
        alt_code = f"A{idx+1:02d}"
        ws_hasil.cell(row=row_num, column=1, value=idx + 1)
        ws_hasil.cell(row=row_num, column=2, value=alt_code)
        ws_hasil.cell(row=row_num, column=3, value=f"='6.Gap'!L{row_num}")
        ws_hasil.cell(row=row_num, column=4, value=f"='6.Gap'!M{row_num}")
        ws_hasil.cell(row=row_num, column=5, value=f'=IF(D{row_num}=1,"★ Peringkat 1 – Sangat Direkomendasikan",IF(D{row_num}<=3,"Sangat Direkomendasikan",IF(D{row_num}<=7,"Direkomendasikan","Pertimbangkan")))')

    # Re-write the BOBOT KRITERIA (SWARA) rekap table header & rows correctly
    ws_hasil.cell(row=27, column=1, value="No")
    ws_hasil.cell(row=27, column=2, value="Kode")
    ws_hasil.cell(row=27, column=3, value="Nama Kriteria")
    ws_hasil.cell(row=27, column=4, value="Tipe")
    ws_hasil.cell(row=27, column=5, value="Bobot Wj")
    
    swara_rekap_data = [
        ['1', 'K1', 'Hasil Proses Wawancara', 'Benefit', "='2.SWARA'!F19"],
        ['2', 'K2', 'Indeks Prestasi Kumulatif', 'Benefit', "='2.SWARA'!F20"],
        ['3', 'K3', 'Status Kepemilikan KIP', 'Benefit', "='2.SWARA'!F21"],
        ['4', 'K4', 'Status Kepemilikan KKS', 'Benefit', "='2.SWARA'!F22"],
        ['5', 'K5', 'Penghasilan Orang Tua (Ayah)', 'Cost', "='2.SWARA'!F23"],
        ['6', 'K6', 'Penghasilan Orang Tua (Ibu)', 'Cost', "='2.SWARA'!F24"],
        ['7', 'K7', 'Status Kepemilikan Rumah', 'Cost', "='2.SWARA'!F25"],
        ['8', 'K8', 'Besaran Daya Listrik', 'Cost', "='2.SWARA'!F26"],
        ['9', 'K9', 'Besaran Luas Tanah', 'Cost', "='2.SWARA'!F27"],
        ['10', 'K10', 'Jenis Sumber Air', 'Cost', "='2.SWARA'!F28"]
    ]
    
    for idx, row_data in enumerate(swara_rekap_data):
        row_num = 28 + idx
        for col_idx, val in enumerate(row_data, 1):
            ws_hasil.cell(row=row_num, column=col_idx, value=val)
        # Copy style from row 28 to make sure layout is neat and zebra formatting is correct
        if row_num > 28:
            copy_row_style(ws_hasil, 28, row_num, 5)

    wb.save("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
    print("Successfully fixed formulas and data in Hasil sheet (including line 31 K4 rekap)!")

if __name__ == "__main__":
    fix_references()
