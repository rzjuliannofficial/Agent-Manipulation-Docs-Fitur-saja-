import openpyxl

wb = openpyxl.load_workbook("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
ws = wb["8.Perbandingan"]

# Let's print out rows 6 to 25 to see what references and formulas we wrote
print("Row | Alt | SAW V | Rank SAW | WP V | Rank WP | TOPSIS V | Rank TOPSIS | MAIRCA Qi | Rank MAIRCA")
for r in range(6, 26):
    alt = ws.cell(row=r, column=1).value
    saw_v = ws.cell(row=r, column=2).value
    saw_r = ws.cell(row=r, column=3).value
    wp_v = ws.cell(row=r, column=4).value
    wp_r = ws.cell(row=r, column=5).value
    top_v = ws.cell(row=r, column=6).value
    top_r = ws.cell(row=r, column=7).value
    ma_qi = ws.cell(row=r, column=8).value
    ma_r = ws.cell(row=r, column=9).value
    print(f"{r:2d}  | {alt} | {saw_v} | {saw_r} | {wp_v} | {wp_r} | {top_v} | {top_r} | {ma_qi} | {ma_r}")

print("\nSpearman Table:")
print("Row | Metode | Sum d^2 | rs formula | hubungan formula")
for r in [31, 32, 33]:
    metode = ws.cell(row=r, column=11).value
    sum_d2 = ws.cell(row=r, column=12).value
    rs = ws.cell(row=r, column=13).value
    hub = ws.cell(row=r, column=14).value
    print(f"{r}  | {metode} | {sum_d2} | {rs} | {hub}")
