import openpyxl

wb = openpyxl.load_workbook("referensi/data_smartphone.xlsx")
ws = wb["1.Data"]

# Print rows 1 to 5, columns 1 to 15
print("--- Rows 1-5 ---")
for r in range(1, 6):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 16)]
    print(f"Row {r}: {row_vals}")

# Print row 9 (first alternative)
print("\n--- Row 9 (Alternative 1) ---")
print([ws.cell(row=9, column=c).value for c in range(1, 16)])
