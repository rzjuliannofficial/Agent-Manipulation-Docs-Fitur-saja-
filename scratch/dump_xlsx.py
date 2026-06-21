import sys
import openpyxl

wb = openpyxl.load_workbook("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=True)
with open("scratch/xlsx_dump.txt", "w", encoding="utf-8") as f:
    f.write("Sheets in workbook: " + str(wb.sheetnames) + "\n")
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        f.write(f"\n=========================================\nSheet: {sheetname}\n")
        f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            # remove trailing Nones
            while row_vals and row_vals[-1] is None:
                row_vals.pop()
            if any(x is not None for x in row_vals):
                f.write(f"Row {r:02d}: {row_vals}\n")
print("Done writing dump.")
