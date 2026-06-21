import openpyxl

wb = openpyxl.load_workbook("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=False)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet: {name}")
    # find some cell with formula
    found = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.startswith("="):
                print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
                found += 1
                if found > 5:
                    break
        if found > 5:
            break
