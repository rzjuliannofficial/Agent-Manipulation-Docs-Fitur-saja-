import openpyxl
from openpyxl.utils import get_column_letter

def revert_excel():
    wb = openpyxl.load_workbook("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
    
    # 1. Revert '1.Data'
    ws_data = wb['1.Data']
    # Check if we have 20 alternatives (row 38 has 'A20' or 20)
    val_38_1 = ws_data.cell(row=38, column=1).value
    val_38_3 = ws_data.cell(row=38, column=3).value
    if val_38_1 == 'A20' or val_38_1 == 20 or val_38_3 == 'A20':
        ws_data.delete_rows(34, 5)
        # Revert MAX/MIN formulas to row 35 and 36
        for col in range(4, 14):
            col_letter = get_column_letter(col)
            ws_data.cell(row=35, column=col, value=f"=MAX({col_letter}19:{col_letter}33)")
            ws_data.cell(row=36, column=col, value=f"=MIN({col_letter}19:{col_letter}33)")
            
    # 2. Revert '3.Normalisasi'
    ws_norm = wb['3.Normalisasi']
    for col in range(2, 12):
        col_letter = get_column_letter(col + 2)
        ws_norm.cell(row=3, column=col, value=f"='1.Data'!{col_letter}35")
        ws_norm.cell(row=4, column=col, value=f"='1.Data'!{col_letter}36")
    if ws_norm.cell(row=25, column=1).value == 'A20':
        ws_norm.delete_rows(21, 5)
        
    # 3. Revert '4.Pref.Teoritis'
    ws_pref_t = wb['4.Pref.Teoritis']
    if ws_pref_t.cell(row=24, column=1).value == 'A20':
        ws_pref_t.delete_rows(20, 5)
    for idx in range(15):
        row_num = 5 + idx
        for col in range(2, 12):
            ws_pref_t.cell(row=row_num, column=col, value=f"='2.SWARA'!F{col+17}/15")
            
    # 4. Revert '5.Pref.Nyata'
    ws_pref_n = wb['5.Pref.Nyata']
    if ws_pref_n.cell(row=23, column=1).value == 'A20':
        ws_pref_n.delete_rows(19, 5)
        
    # 5. Revert '6.Gap'
    ws_gap = wb['6.Gap']
    if ws_gap.cell(row=23, column=1).value == 'A20':
        ws_gap.delete_rows(19, 5)
    for idx in range(15):
        row_num = 4 + idx
        ws_gap.cell(row=row_num, column=13, value=f"=RANK(L{row_num},$L$4:$L$18,1)")
        
    # 6. Revert '7.Hasil'
    ws_hasil = wb['7.Hasil']
    if ws_hasil.cell(row=23, column=2).value == 'A20':
        ws_hasil.delete_rows(19, 5)
        
    # Revert merged cell in Hasil from A26:E26 to A21:E21
    is_a26_merged = False
    for r in list(ws_hasil.merged_cells.ranges):
        if r.coord == "A26:E26":
            is_a26_merged = True
            break
    if is_a26_merged:
        ws_hasil.unmerge_cells("A26:E26")
        ws_hasil['A26'] = None
        
        is_a21_merged = False
        for r in list(ws_hasil.merged_cells.ranges):
            if r.coord == "A21:E21":
                is_a21_merged = True
                break
        if not is_a21_merged:
            ws_hasil.merge_cells("A21:E21")
        ws_hasil['A21'] = "BOBOT KRITERIA (SWARA) – REKAP AKHIR"
        
    # Fix SWARA weights formulas in Hasil
    for idx in range(10):
        row_num = 22 + idx
        ws_hasil.cell(row=row_num, column=5, value=f"='2.SWARA'!F{idx+19}")
        
    if '8.Perbandingan' in wb.sheetnames:
        del wb['8.Perbandingan']
        
    wb.save("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
    print("Workbook successfully reverted to 15 alternatives.")

if __name__ == "__main__":
    revert_excel()
