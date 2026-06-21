import openpyxl

wb = openpyxl.load_workbook("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=True)
ws = wb["2.SWARA"]

print("--- SWARA Sheet Rows 4-15 ---")
for r in range(4, 16):
    prio = ws.cell(row=r, column=1).value
    code = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    sj = ws.cell(row=r, column=4).value
    kj = ws.cell(row=r, column=5).value
    qj = ws.cell(row=r, column=6).value
    wj = ws.cell(row=r, column=7).value
    print(f"Prio {prio} | Code {code} | Name {name} | Sj {sj} | Kj {kj} | Qj {qj} | Wj {wj}")
    
# Sum Qj
sum_qj = 0
for r in range(4, 16):
    sum_qj += float(ws.cell(row=r, column=6).value or 0.0)
print(f"Sum Qj: {sum_qj}")
