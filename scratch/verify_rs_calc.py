import openpyxl

wb = openpyxl.load_workbook("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
ws_data = wb["1.Data"]

c_codes = []
c_types = []
col = 4
while True:
    val = ws_data.cell(row=2, column=col).value
    if not val:
        break
    c_codes.append(val)
    c_types.append(ws_data.cell(row=3, column=col).value)
    col += 1
num_criteria = len(c_codes)

alt_codes = []
alt_data = []
row_idx = 9
while True:
    val = ws_data.cell(row=row_idx, column=1).value
    if not val:
        break
    alt_codes.append(val)
    row_vals = []
    for c_idx in range(num_criteria):
        row_vals.append(float(ws_data.cell(row=row_idx, column=c_idx + 2).value or 0.0))
    alt_data.append(row_vals)
    row_idx += 1
num_alt = len(alt_codes)

ws_swara = wb["2.SWARA"]
swara_rows = []
for r_idx in range(num_criteria):
    sj = float(ws_swara.cell(row=4 + r_idx, column=4).value or 0.0)
    swara_rows.append({"sj": sj})

swara_rows[0]["kj"] = 1.0
swara_rows[0]["qj"] = 1.0
for i in range(1, len(swara_rows)):
    swara_rows[i]["kj"] = swara_rows[i]["sj"] + 1.0
    swara_rows[i]["qj"] = swara_rows[i - 1]["qj"] / swara_rows[i]["kj"]
    
sum_qj = sum(item["qj"] for item in swara_rows)
for item in swara_rows:
    item["wj"] = item["qj"] / sum_qj
    
ordered_weights = [item["wj"] for item in swara_rows]

# MAIRCA
max_vals = [max(alt_data[r][c] for r in range(num_alt)) for c in range(num_criteria)]
min_vals = [min(alt_data[r][c] for r in range(num_alt)) for c in range(num_criteria)]

norm_data = []
for r_idx in range(num_alt):
    row_norm = []
    for c_idx in range(num_criteria):
        is_benefit = c_types[c_idx].strip().lower() == "benefit"
        x_val = alt_data[r_idx][c_idx]
        mn = min_vals[c_idx]
        mx = max_vals[c_idx]
        n_val = (x_val - mn) / (mx - mn) if mx != mn else 1.0
        if not is_benefit:
            n_val = 1.0 - n_val
        row_norm.append(n_val)
    norm_data.append(row_norm)
    
tp_row = [wj / num_alt for wj in ordered_weights]

qi_list = []
for r_idx in range(num_alt):
    row_gap_sum = 0.0
    for c_idx in range(num_criteria):
        tp_val = tp_row[c_idx]
        pp_val = tp_val * norm_data[r_idx][c_idx]
        g_val = tp_val - pp_val
        row_gap_sum += g_val
    qi_list.append({"alt": alt_codes[r_idx], "qi": row_gap_sum})
    
sorted_qi = sorted(qi_list, key=lambda x: x["qi"])
ma_ranks = {}
for rank, item in enumerate(sorted_qi, start=1):
    ma_ranks[item["alt"]] = rank

# SAW, WP, TOPSIS
def calculate_saw_wp_topsis(alt_data, c_types, weights):
    num_alt = len(alt_data)
    num_crit = len(c_types)
    saw_v = []
    for r in range(num_alt):
        val = 0.0
        for c in range(num_crit):
            is_benefit = c_types[c].strip().lower() == "benefit"
            if is_benefit:
                norm = alt_data[r][c] / max_vals[c] if max_vals[c] != 0 else 0.0
            else:
                norm = min_vals[c] / alt_data[r][c] if alt_data[r][c] != 0 else 0.0
            val += norm * weights[c]
        saw_v.append(val)
    wp_v = []
    for r in range(num_alt):
        val = 1.0
        for c in range(num_crit):
            is_benefit = c_types[c].strip().lower() == "benefit"
            power = weights[c] if is_benefit else -weights[c]
            base_val = alt_data[r][c] if alt_data[r][c] != 0 else 1e-9
            val *= (base_val ** power)
        wp_v.append(val)
    dividers = []
    for c in range(num_crit):
        sum_sq = sum(alt_data[r][c]**2 for r in range(num_alt))
        dividers.append(sum_sq**0.5)
    v_matrix = []
    for r in range(num_alt):
        row_v = []
        for c in range(num_crit):
            norm = alt_data[r][c] / dividers[c] if dividers[c] != 0 else 0.0
            row_v.append(norm * weights[c])
        v_matrix.append(row_v)
    a_plus = []
    a_minus = []
    for c in range(num_crit):
        col_v = [v_matrix[r][c] for r in range(num_alt)]
        is_benefit = c_types[c].strip().lower() == "benefit"
        if is_benefit:
            a_plus.append(max(col_v))
            a_minus.append(min(col_v))
        else:
            a_plus.append(min(col_v))
            a_minus.append(max(col_v))
    topsis_v = []
    for r in range(num_alt):
        d_plus = sum((v_matrix[r][c] - a_plus[c])**2 for c in range(num_crit))**0.5
        d_minus = sum((v_matrix[r][c] - a_minus[c])**2 for c in range(num_crit))**0.5
        val = d_minus / (d_plus + d_minus) if (d_plus + d_minus) != 0 else 0.0
        topsis_v.append(val)
    return saw_v, wp_v, topsis_v

saw_v, wp_v, topsis_v = calculate_saw_wp_topsis(alt_data, c_types, ordered_weights)

def get_ranks(values, reverse=True):
    sorted_vals = sorted(enumerate(values), key=lambda x: x[1], reverse=reverse)
    ranks = [0] * len(values)
    for rank, (idx, val) in enumerate(sorted_vals, start=1):
        ranks[idx] = rank
    return ranks

saw_ranks = get_ranks(saw_v, True)
wp_ranks = get_ranks(wp_v, True)
topsis_ranks = get_ranks(topsis_v, True)
mairca_ranks_aligned = [ma_ranks[alt] for alt in alt_codes]

def spearman_rs(ranks1, ranks2):
    sum_d2 = sum((r1 - r2) ** 2 for r1, r2 in zip(ranks1, ranks2))
    n = len(ranks1)
    rs = 1.0 - (6.0 * sum_d2) / (n * (n**2 - 1))
    return sum_d2, rs

d2_saw, rs_saw = spearman_rs(mairca_ranks_aligned, saw_ranks)
d2_wp, rs_wp = spearman_rs(mairca_ranks_aligned, wp_ranks)
d2_topsis, rs_topsis = spearman_rs(mairca_ranks_aligned, topsis_ranks)

print(f"MAIRCA vs SAW: sum_d2={d2_saw}, rs={rs_saw:.4f}")
print(f"MAIRCA vs WP: sum_d2={d2_wp}, rs={rs_wp:.4f}")
print(f"MAIRCA vs TOPSIS: sum_d2={d2_topsis}, rs={rs_topsis:.4f}")
