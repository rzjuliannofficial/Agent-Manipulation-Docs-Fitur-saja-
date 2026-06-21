import openpyxl

wb = openpyxl.load_workbook("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=True)
ws = wb["1.Data"]

# Print criteria details
# K1: Weight g (Cost)
# K3: Year of production (Benefit)

k1_vals = [float(ws.cell(row=r, column=2).value) for r in range(9, 29)]
k3_vals = [float(ws.cell(row=r, column=4).value) for r in range(9, 29)]

print("K1 (Weight g):")
print(f"  A01: {ws.cell(row=9, column=2).value}")
print(f"  Min: {min(k1_vals)}")
print(f"  Max: {max(k1_vals)}")

print("\nK3 (Year of production):")
print(f"  A01: {ws.cell(row=9, column=4).value}")
print(f"  Min: {min(k3_vals)}")
print(f"  Max: {max(k3_vals)}")
