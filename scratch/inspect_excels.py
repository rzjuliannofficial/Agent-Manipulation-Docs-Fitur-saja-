import openpyxl

def inspect_excel(path):
    print(f"\n=== Inspecting {path} ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    
    # 2. SWARA weights
    ws_swara = wb["2.SWARA"]
    print("SWARA weights (first 3 and last 1):")
    for r in range(4, 16):
        code = ws_swara.cell(row=r, column=2).value
        weight = ws_swara.cell(row=r, column=7).value
        # If weight is None, check column 5 of Rekap
        # Let's print out what is there
        print(f"  Row {r}: Code={code}, Weight={weight}")
        
    print("\nRekap weights in SWARA:")
    # Rekap starts at 4 + 12 + 4 = 20
    for r in range(20, 32):
        code = ws_swara.cell(row=r, column=2).value
        weight = ws_swara.cell(row=r, column=5).value
        print(f"  Row {r}: Code={code}, Weight={weight}")
        
    # 7. Hasil Rankings
    ws_hasil = wb["7.Hasil"]
    print("\nHasil Rankings (first 3):")
    for r in range(5, 8):
        alt = ws_hasil.cell(row=r, column=2).value
        qi = ws_hasil.cell(row=r, column=3).value
        rank = ws_hasil.cell(row=r, column=4).value
        print(f"  Rank {r-4}: Alt={alt}, Qi={qi}, Rank={rank}")
        
inspect_excel("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
inspect_excel("output/SWARA_MAIRCA_DETAIL_RUMUS (1).xlsx")
