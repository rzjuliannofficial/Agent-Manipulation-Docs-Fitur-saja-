import openpyxl

def build_data():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Data
    ws_data = wb.active
    ws_data.title = "1.Data"
    
    ws_data["A1"] = "MATRIKS DATA KEPUTUSAN — SELEKSI SMARTPHONE TERBAIK"
    ws_data["A2"] = "Keterangan Kriteria"
    ws_data["A3"] = "Tipe Kriteria"
    ws_data["A4"] = "Nama Kriteria"
    
    # Headers A2 to M4
    headers = [
        ("K1", "Cost", "Weight g"),
        ("K2", "Cost", "Weight oz"),
        ("K3", "Benefit", "Year of production"),
        ("K4", "Benefit", "Display resolution"),
        ("K5", "Benefit", "Memory card"),
        ("K6", "Benefit", "Internal memory"),
        ("K7", "Benefit", "RAM"),
        ("K8", "Benefit", "Primary camera"),
        ("K9", "Benefit", "Secondary camera"),
        ("K10", "Benefit", "Bluetooth"),
        ("K11", "Benefit", "USB"),
        ("K12", "Cost", "Approx price"),
    ]
    for i, (k_code, k_type, k_name) in enumerate(headers, start=4):
        ws_data.cell(row=2, column=i, value=k_code)
        ws_data.cell(row=3, column=i, value=k_type)
        ws_data.cell(row=4, column=i, value=k_name)
        
    ws_data["A7"] = "MATRIKS KEPUTUSAN (Nilai Atribut tiap Alternatif)"
    ws_data["A8"] = "Alternatif"
    for i in range(1, 13):
        ws_data.cell(row=8, column=i+1, value=f"K{i}")
        
    # Alternatives data
    alternatives = [
        ("A01", [202, 7, 2015, 5, 64, 8, 2048, 13, 5, 4, 2, 170]),
        ("A02", [150, 5, 2014, 5, 32, 4, 1024, 8, 2, 4, 2, 180]),
        ("A03", [160, 5, 2016, 5, 256, 32, 3072, 16, 8, 4, 1, 420]),
        ("A04", [170, 6, 2016, 5, 32, 16, 2048, 8, 8, 4, 2, 110]),
        ("A05", [115, 6, 2015, 4, 64, 8, 1024, 8, 5, 4, 2, 180]),
        ("A06", [151, 5, 2016, 5, 256, 64, 4096, 16, 8, 4, 2, 440]),
        ("A07", [310, 10, 2017, 8, 256, 16, 2048, 5, 2, 4, 2, 100]),
        ("A08", [145, 11, 2015, 4, 32, 8, 1024, 5, 2, 4, 2, 130]),
        ("A09", [110, 3, 2015, 5, 32, 8, 1024, 13, 5, 4, 2, 200]),
        ("A10", [160, 5, 2016, 5, 256, 32, 3072, 16, 8, 4, 1, 420]),
        ("A11", [130, 4, 2013, 4, 64, 8, 1024, 8, 1, 4, 2, 290]),
        ("A12", [170, 6, 2015, 5, 256, 16, 2048, 13, 5, 4, 2, 370]),
        ("A13", [143, 4, 2017, 5, 256, 16, 2048, 13, 5, 4, 2, 150]),
        ("A14", [176, 6, 2013, 5, 32, 4, 1024, 8, 1, 4, 2, 230]),
        ("A15", [328, 11, 2014, 8, 64, 8, 1024, 5, 2, 4, 2, 200]),
        ("A16", [132, 4, 2012, 4, 32, 8, 1024, 8, 1, 3, 2, 210]),
        ("A17", [135, 4, 2016, 5, 256, 16, 2048, 13, 8, 4, 2, 250]),
        ("A18", [170, 6, 2011, 4, 32, 16, 1024, 8, 2, 3, 2, 160]),
        ("A19", [150, 5, 2014, 5, 32, 8, 1024, 13, 5, 4, 2, 190]),
        ("A20", [143, 4, 2017, 5, 256, 16, 2048, 13, 5, 4, 2, 150]),
    ]
    for r_idx, (alt_code, alt_vals) in enumerate(alternatives, start=9):
        ws_data.cell(row=r_idx, column=1, value=alt_code)
        for c_idx, val in enumerate(alt_vals, start=2):
            ws_data.cell(row=r_idx, column=c_idx, value=val)
            
    # Sheet 2: SWARA
    ws_swara = wb.create_sheet(title="2.SWARA")
    ws_swara["A1"] = "PEMBOBOTAN KRITERIA — METODE SWARA"
    ws_swara["A3"] = "Prioritas"
    ws_swara["B3"] = "Kode"
    ws_swara["C3"] = "Nama Kriteria"
    ws_swara["D3"] = "Sj (input)"
    
    # Priority order and Sj input
    swara_inputs = [
        (1, "K12", "Approx price", 0.0),
        (2, "K7", "RAM", 0.05),
        (3, "K6", "Internal memory", 0.05),
        (4, "K3", "Year of production", 0.10),
        (5, "K8", "Primary camera", 0.10),
        (6, "K5", "Memory card", 0.15),
        (7, "K4", "Display resolution", 0.10),
        (8, "K9", "Secondary camera", 0.15),
        (9, "K1", "Weight g", 0.15),
        (10, "K10", "Bluetooth", 0.10),
        (11, "K11", "USB", 0.10),
        (12, "K2", "Weight oz", 0.05),
    ]
    for r_idx, (prio, k_code, k_name, sj) in enumerate(swara_inputs, start=4):
        ws_swara.cell(row=r_idx, column=1, value=prio)
        ws_swara.cell(row=r_idx, column=2, value=k_code)
        ws_swara.cell(row=r_idx, column=3, value=k_name)
        ws_swara.cell(row=r_idx, column=4, value=sj)
        
    wb.save("referensi/data_smartphone.xlsx")
    print("Successfully built referensi/data_smartphone.xlsx")

if __name__ == "__main__":
    build_data()
