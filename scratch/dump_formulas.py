import openpyxl

wb = openpyxl.load_workbook("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=False)
with open("scratch/xlsx_formulas.txt", "w", encoding="utf-8") as f:
    for name in wb.sheetnames:
        ws = wb[name]
        f.write(f"\n=========================================\nSheet: {name}\n")
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and val.startswith("="):
                    col_letter = openpyxl.utils.get_column_letter(c)
                    f.write(f"Cell {col_letter}{r}: {val}\n")
print("Done writing formulas.")
