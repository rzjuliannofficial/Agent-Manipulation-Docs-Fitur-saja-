import openpyxl

wb = openpyxl.load_workbook("output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
ws_data = wb["1.Data"]

c_codes = []
c_types = []
col = 4
while True:
    val = ws_data.cell(row=2, column=col).value
    if not val:
        break
    c_codes.append(val)
    c_types.append(ws_data.cell(row=3, column=col).value)
    col += 1
num_criteria = len(c_codes)

alt_codes = []
alt_data = []
row_idx = 9
while True:
    val = ws_data.cell(row=row_idx, column=1).value
    if not val:
        break
    alt_codes.append(val)
    row_vals = []
    for c_idx in range(num_criteria):
        row_vals.append(float(ws_data.cell(row=row_idx, column=c_idx + 2).value or 0.0))
    alt_data.append(row_vals)
    row_idx += 1
num_alt = len(alt_codes)

ws_swara = wb["2.SWARA"]
swara_rows = []
for r_idx in range(num_criteria):
    code = ws_swara.cell(row=4 + r_idx, column=2).value
    sj = float(ws_swara.cell(row=4 + r_idx, column=4).value or 0.0)
    swara_rows.append({"code": code, "sj": sj})

swara_rows[0]["kj"] = 1.0
swara_rows[0]["qj"] = 1.0
for i in range(1, len(swara_rows)):
    swara_rows[i]["kj"] = swara_rows[i]["sj"] + 1.0
    swara_rows[i]["qj"] = swara_rows[i - 1]["qj"] / swara_rows[i]["kj"]
    
sum_qj = sum(item["qj"] for item in swara_rows)
for item in swara_rows:
    item["wj"] = item["qj"] / sum_qj
    
weights_map = {item["code"]: item["wj"] for item in swara_rows}
ordered_weights = [weights_map[code] for code in c_codes]

# MAIRCA
max_vals = [max(alt_data[r][c] for r in range(num_alt)) for c in range(num_criteria)]
min_vals = [min(alt_data[r][c] for r in range(num_alt)) for c in range(num_criteria)]

norm_data = []
for r_idx in range(num_alt):
    row_norm = []
    for c_idx in range(num_criteria):
        is_benefit = c_types[c_idx].strip().lower() == "benefit"
        x_val = alt_data[r_idx][c_idx]
        mn = min_vals[c_idx]
        mx = max_vals[c_idx]
        n_val = (x_val - mn) / (mx - mn) if mx != mn else 1.0
        if not is_benefit:
            n_val = 1.0 - n_val
        row_norm.append(n_val)
    norm_data.append(row_norm)
    
tp_row = [wj / num_alt for wj in ordered_weights]

# Find A06 and A01
a06_idx = alt_codes.index("A06")
a01_idx = alt_codes.index("A01")
k12_idx = c_codes.index("K12")
k2_idx = c_codes.index("K2")

norm_a06_k12 = norm_data[a06_idx][k12_idx]
norm_a01_k2 = norm_data[a01_idx][k2_idx]

tp_k12 = tp_row[k12_idx]
tp_k2 = tp_row[k2_idx]

pp_a06_k12 = tp_k12 * norm_a06_k12
pp_a01_k2 = tp_k2 * norm_a01_k2

print(f"Norm A06 K12: {norm_a06_k12:.4f}")
print(f"Norm A01 K2: {norm_a01_k2:.4f}")
print(f"TP K12: {tp_k12:.5f}")
print(f"TP K2: {tp_k2:.5f}")
print(f"PP A06 K12: {pp_a06_k12:.5f}")
print(f"PP A01 K2: {pp_a01_k2:.5f}")
