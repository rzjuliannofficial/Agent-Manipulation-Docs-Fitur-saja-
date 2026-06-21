import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

def copy_cell_style(src_cell, dest_cell):
    if src_cell.font:
        dest_cell.font = copy.copy(src_cell.font)
    if src_cell.border:
        dest_cell.border = copy.copy(src_cell.border)
    if src_cell.fill:
        dest_cell.fill = copy.copy(src_cell.fill)
    if src_cell.alignment:
        dest_cell.alignment = copy.copy(src_cell.alignment)
    dest_cell.number_format = src_cell.number_format

def copy_row_style(ws, src_row, dest_row, num_cols):
    for col in range(1, num_cols + 1):
        copy_cell_style(ws.cell(row=src_row, column=col), ws.cell(row=dest_row, column=col))

def update_lazizmu_excel():
    # Load the workbook
    wb = openpyxl.load_workbook("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx", data_only=False)
    
    # New alternatives data
    a16_a20_data = [
        ['A16', 4, 4, 1, 1, 5, 1, 1, 1, 3, 3],
        ['A17', 4, 5, 1, 1, 1, 3, 4, 2, 3, 2],
        ['A18', 4, 4, 1, 1, 3, 1, 4, 2, 4, 2],
        ['A19', 4, 3, 1, 1, 5, 1, 4, 1, 1, 2],
        ['A20', 4, 4, 1, 1, 3, 4, 4, 1, 5, 2]
    ]
    
    # -------------------------------------------------------------
    # 1. Update '1.Data'
    # -------------------------------------------------------------
    ws_data = wb['1.Data']
    
    # 1a. Format headers (D2:M4) to match the image exactly
    font_family = "Arial"
    fill_brown = PatternFill(start_color="FF8A5E32", end_color="FF8A5E32", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    
    font_k = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_benefit = Font(name=font_family, size=10, bold=True, color="FF006600") # Dark Green
    font_cost = Font(name=font_family, size=10, bold=True, color="FF990000") # Dark Red
    font_name = Font(name=font_family, size=9, italic=True, color="000000")
    
    border_thin = Side(style='thin', color='A0A0A0')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Row 2: K1-K10
    for idx, k_num in enumerate(range(1, 11)):
        col = 4 + idx
        cell = ws_data.cell(row=2, column=col, value=f"K{k_num}")
        cell.font = font_k
        cell.fill = fill_brown
        cell.alignment = align_center_wrap
        cell.border = cell_border
        
    # Row 3: Benefit / Cost
    types = ["Benefit", "Benefit", "Benefit", "Benefit", "Cost", "Cost", "Cost", "Cost", "Cost", "Cost"]
    for idx, t_val in enumerate(types):
        col = 4 + idx
        cell = ws_data.cell(row=3, column=col, value=t_val)
        cell.font = font_benefit if t_val == "Benefit" else font_cost
        cell.fill = fill_white
        cell.alignment = align_center_wrap
        cell.border = cell_border
        
    # Row 4: Criteria Names
    names = [
        "Hasil Wawancara", "IPK", "Kepemilikan KIP", "Kepemilikan KKS",
        "Penghasilan Ayah", "Penghasilan Ibu", "Status Rumah", "Daya Listrik",
        "Luas Tanah", "Sumber Air"
    ]
    for idx, name in enumerate(names):
        col = 4 + idx
        cell = ws_data.cell(row=4, column=col, value=name)
        cell.font = font_name
        cell.fill = fill_white
        cell.alignment = align_center_wrap
        cell.border = cell_border

    # 1b. Write A16-A20 data rows
    has_a16 = False
    for r in range(19, 45):
        if ws_data.cell(row=r, column=1).value == 'A16':
            has_a16 = True
            break
            
    if not has_a16:
        ws_data.insert_rows(34, 5)
        
    for idx, row_data in enumerate(a16_a20_data):
        row_num = 34 + idx
        ws_data.cell(row=row_num, column=1, value=row_data[0])
        for col_idx, val in enumerate(row_data[1:], 2):
            ws_data.cell(row=row_num, column=col_idx, value=val)
            
        src_row = 19 if row_num % 2 == 1 else 20
        copy_row_style(ws_data, src_row, row_num, 11)
        
    # Fix the MAX and MIN formulas (now at rows 40 and 41)
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        ws_data.cell(row=40, column=col, value=f"=MAX({col_letter}19:{col_letter}38)")
        ws_data.cell(row=41, column=col, value=f"=MIN({col_letter}19:{col_letter}38)")
        
    # -------------------------------------------------------------
    # 3. Update '3.Normalisasi'
    # -------------------------------------------------------------
    ws_norm = wb['3.Normalisasi']
    for col in range(2, 12):
        col_letter = get_column_letter(col + 2)
        ws_norm.cell(row=3, column=col, value=f"='1.Data'!{col_letter}40")
        ws_norm.cell(row=4, column=col, value=f"='1.Data'!{col_letter}41")
        
    has_a16_norm = False
    for r in range(6, 30):
        if ws_norm.cell(row=r, column=1).value == 'A16':
            has_a16_norm = True
            break
            
    if not has_a16_norm:
        ws_norm.insert_rows(21, 5)
        
    for idx in range(20):
        row_num = 6 + idx
        alt_code = f"A{idx+1:02d}"
        ws_norm.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            data_col_letter = get_column_letter(col)
            data_row = 19 + idx
            if col in [2, 3, 4, 5]: # Benefit
                ws_norm.cell(row=row_num, column=col, value=f"=IF(${col_letter}$3=${col_letter}$4,0,('1.Data'!{data_col_letter}{data_row}-${col_letter}$4)/(${col_letter}$3-${col_letter}$4))")
            else: # Cost
                ws_norm.cell(row=row_num, column=col, value=f"=IF(${col_letter}$3=${col_letter}$4,0,(${col_letter}$3-'1.Data'!{data_col_letter}{data_row})/(${col_letter}$3-${col_letter}$4))")
                
        src_row = 6 if row_num % 2 == 0 else 7
        copy_row_style(ws_norm, src_row, row_num, 11)

    # -------------------------------------------------------------
    # 4. Update '4.Pref.Teoritis'
    # -------------------------------------------------------------
    ws_pref_t = wb['4.Pref.Teoritis']
    has_a16_pref_t = False
    for r in range(5, 30):
        if ws_pref_t.cell(row=r, column=1).value == 'A16':
            has_a16_pref_t = True
            break
            
    if not has_a16_pref_t:
        ws_pref_t.insert_rows(20, 5)
        
    for idx in range(20):
        row_num = 5 + idx
        alt_code = f"A{idx+1:02d}"
        ws_pref_t.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            ws_pref_t.cell(row=row_num, column=col, value=f"='2.SWARA'!F{col+17}/20")
            
        src_row = 5 if row_num % 2 == 1 else 6
        copy_row_style(ws_pref_t, src_row, row_num, 11)

    # -------------------------------------------------------------
    # 5. Update '5.Pref.Nyata'
    # -------------------------------------------------------------
    ws_pref_n = wb['5.Pref.Nyata']
    has_a16_pref_n = False
    for r in range(4, 30):
        if ws_pref_n.cell(row=r, column=1).value == 'A16':
            has_a16_pref_n = True
            break
            
    if not has_a16_pref_n:
        ws_pref_n.insert_rows(19, 5)
        
    for idx in range(20):
        row_num = 4 + idx
        alt_code = f"A{idx+1:02d}"
        ws_pref_n.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            pref_t_row = 5 + idx
            norm_row = 6 + idx
            ws_pref_n.cell(row=row_num, column=col, value=f"='4.Pref.Teoritis'!{col_letter}{pref_t_row}*'3.Normalisasi'!{col_letter}{norm_row}")
            
        src_row = 4 if row_num % 2 == 0 else 5
        copy_row_style(ws_pref_n, src_row, row_num, 11)

    # -------------------------------------------------------------
    # 6. Update '6.Gap'
    # -------------------------------------------------------------
    ws_gap = wb['6.Gap']
    has_a16_gap = False
    for r in range(4, 30):
        if ws_gap.cell(row=r, column=1).value == 'A16':
            has_a16_gap = True
            break
            
    if not has_a16_gap:
        ws_gap.insert_rows(19, 5)
        
    for idx in range(20):
        row_num = 4 + idx
        alt_code = f"A{idx+1:02d}"
        ws_gap.cell(row=row_num, column=1, value=alt_code)
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            pref_t_row = 5 + idx
            pref_n_row = 4 + idx
            ws_gap.cell(row=row_num, column=col, value=f"='4.Pref.Teoritis'!{col_letter}{pref_t_row}-'5.Pref.Nyata'!{col_letter}{pref_n_row}")
        ws_gap.cell(row=row_num, column=12, value=f"=SUM(B{row_num}:K{row_num})")
        ws_gap.cell(row=row_num, column=13, value=f"=RANK(L{row_num},$L$4:$L$23,1)")
        
        src_row = 4 if row_num % 2 == 0 else 5
        copy_row_style(ws_gap, src_row, row_num, 13)

    # -------------------------------------------------------------
    # 7. Update '7.Hasil'
    # -------------------------------------------------------------
    ws_hasil = wb['7.Hasil']
    has_a16_hasil = False
    for r in range(4, 30):
        if ws_hasil.cell(row=r, column=2).value == 'A16':
            has_a16_hasil = True
            break
            
    if not has_a16_hasil:
        ws_hasil.unmerge_cells("A1:E1")
        ws_hasil.unmerge_cells("A2:E2")
        ws_hasil.unmerge_cells("A21:E21")
        
        ws_hasil.insert_rows(19, 5)
        
        ws_hasil.merge_cells("A1:E1")
        ws_hasil.merge_cells("A2:E2")
        ws_hasil.merge_cells("A26:E26")
        
        ws_hasil['A26'] = "BOBOT KRITERIA (SWARA) – REKAP AKHIR"
        copy_row_style(ws_hasil, 3, 26, 5)

    for idx in range(20):
        row_num = 4 + idx
        alt_code = f"A{idx+1:02d}"
        ws_hasil.cell(row=row_num, column=1, value=idx + 1)
        ws_hasil.cell(row=row_num, column=2, value=alt_code)
        ws_hasil.cell(row=row_num, column=3, value=f"='6.Gap'!L{row_num}")
        ws_hasil.cell(row=row_num, column=4, value=f"='6.Gap'!M{row_num}")
        ws_hasil.cell(row=row_num, column=5, value=f'=IF(D{row_num}=1,"★ Peringkat 1 – Sangat Direkomendasikan",IF(D{row_num}<=3,"Sangat Direkomendasikan",IF(D{row_num}<=7,"Direkomendasikan","Pertimbangkan")))')
        
        src_row = 4 if row_num % 2 == 0 else 5
        copy_row_style(ws_hasil, src_row, row_num, 5)

    # Fix SWARA weight rekap rows in Hasil
    for idx in range(10):
        row_num = 27 + idx
        ws_hasil.cell(row=row_num, column=5, value=f"='2.SWARA'!F{idx+19}")

    # -------------------------------------------------------------
    # 8. Create and populate '8.Perbandingan' Sheet
    # -------------------------------------------------------------
    if '8.Perbandingan' in wb.sheetnames:
        del wb['8.Perbandingan']
    ws_comp = wb.create_sheet('8.Perbandingan')
    ws_comp.views.sheetView[0].showGridLines = True
    
    font_family = "Arial"
    
    font_title = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="000000")
    font_section = Font(name=font_family, size=10, bold=True, color="000000")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10, color="000000")
    font_bold_data = Font(name=font_family, size=10, bold=True, color="000000")
    
    fill_title = PatternFill(start_color="FF404040", end_color="FF404040", fill_type="solid")
    fill_header = PatternFill(start_color="FF737373", end_color="FF737373", fill_type="solid")
    fill_zebra = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    fill_highlight = PatternFill(start_color="FFE6E6E6", end_color="FFE6E6E6", fill_type="solid")
    
    border_thin = Side(style='thin', color='A0A0A0')
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Title Block
    ws_comp.row_dimensions[1].height = 25
    ws_comp.merge_cells("A1:I1")
    ws_comp["A1"] = "ANALISIS PERBANDINGAN METODE DAN UJI SENSITIVITAS"
    ws_comp["A1"].font = font_title
    ws_comp["A1"].fill = fill_title
    ws_comp["A1"].alignment = align_center
    
    ws_comp.row_dimensions[2].height = 20
    ws_comp.merge_cells("A2:I2")
    ws_comp["A2"] = "Berdasarkan Data Seleksi Beasiswa Lazizmu (20 Alternatif)"
    ws_comp["A2"].font = font_subtitle
    ws_comp["A2"].alignment = align_center
    
    # Table 1 Title
    ws_comp["A4"] = "Tabel 1. Perbandingan Nilai dan Peringkat SAW, WP, TOPSIS, dan SWARA-MAIRCA"
    ws_comp["A4"].font = font_section
    ws_comp["A4"].alignment = align_left
    
    headers = [
        "Alternatif", 
        "SAW V", "Rank SAW", 
        "WP V", "Rank WP", 
        "TOPSIS V", "Rank TOPSIS", 
        "MAIRCA Qi", "Rank MAIRCA"
    ]
    
    ws_comp.row_dimensions[5].height = 24
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws_comp.cell(row=5, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border
        
    saw_v_values = [16.80, 13.77, 17.32, 12.00, 15.22, 12.60, 13.85, 13.95, 13.87, 13.35, 
                    12.75, 14.28, 11.75, 16.95, 11.02, 13.27, 13.05, 12.37, 12.75, 11.32]
    
    wp_v_values = [66.806, 49.833, 69.343, 42.512, 59.856, 43.040, 50.127, 53.380, 50.958, 48.420, 
                   45.563, 54.086, 41.414, 66.550, 39.432, 45.858, 45.897, 43.752, 44.180, 38.994]
    
    topsis_v_values = [0.728656, 0.49106, 0.765337, 0.460397, 0.65903, 0.458221, 0.518355, 0.594643, 0.558432, 0.497844, 
                       0.474355, 0.548282, 0.443418, 0.7578, 0.336822, 0.466242, 0.470457, 0.488074, 0.442941, 0.329795]
    
    for i in range(20):
        row_num = 6 + i
        ws_comp.row_dimensions[row_num].height = 20
        alt_code = f"A{i+1:02d}"
        
        ws_comp.cell(row=row_num, column=1, value=alt_code).alignment = align_center
        
        # SAW
        ws_comp.cell(row=row_num, column=2, value=saw_v_values[i]).number_format = '0.00'
        ws_comp.cell(row=row_num, column=3, value=f"=RANK(B{row_num},$B$6:$B$25,0)")
        
        # WP
        ws_comp.cell(row=row_num, column=4, value=wp_v_values[i]).number_format = '0.000'
        ws_comp.cell(row=row_num, column=5, value=f"=RANK(D{row_num},$D$6:$D$25,0)")
        
        # TOPSIS
        ws_comp.cell(row=row_num, column=6, value=topsis_v_values[i]).number_format = '0.000000'
        ws_comp.cell(row=row_num, column=7, value=f"=RANK(F{row_num},$F$6:$F$25,0)")
        
        # SWARA-MAIRCA Qi
        ws_comp.cell(row=row_num, column=8, value=f"='7.Hasil'!C{i+4}").number_format = '0.000000'
        ws_comp.cell(row=row_num, column=9, value=f"='7.Hasil'!D{i+4}")
        
        is_even = i % 2 == 1
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)
        
        for c in range(1, 10):
            cell = ws_comp.cell(row=row_num, column=c)
            cell.font = font_bold_data if c in [3, 5, 7, 9] else font_data
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = box_border
            if c != 1:
                cell.alignment = align_right if c in [2, 4, 6, 8] else align_center

    # Summary Text Box below Table 1
    sum_row = 27
    ws_comp.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=9)
    sum_cell = ws_comp.cell(row=sum_row, column=1, value="*Alternatif A03 menempati Peringkat 1 di seluruh metode pembanding utama (SAW, WP, TOPSIS, dan SWARA-MAIRCA)")
    sum_cell.font = Font(name=font_family, size=9, bold=True, italic=True, color="000000")
    sum_cell.alignment = align_left
    
    # Table 2: Spearman Correlation
    ws_comp["K4"] = "Tabel 2. Perhitungan Korelasi Spearman"
    ws_comp["K4"].font = font_section
    ws_comp["K4"].alignment = align_left
    
    comp_headers = [
        "Alt", 
        "Rank R1\n(MAIRCA)", 
        "Rank R2\n(SAW)", "d² (SAW)", 
        "Rank R3\n(WP)", "d² (WP)", 
        "Rank R4\n(TOPSIS)", "d² (TOPSIS)"
    ]
    
    ws_comp.row_dimensions[5].height = 24
    for col_idx, h_text in enumerate(comp_headers, 11):
        cell = ws_comp.cell(row=5, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border
        
    for i in range(20):
        row_num = 6 + i
        alt_code = f"A{i+1:02d}"
        
        ws_comp.cell(row=row_num, column=11, value=alt_code).alignment = align_center
        ws_comp.cell(row=row_num, column=12, value=f"=I{row_num}").alignment = align_center
        
        # SAW
        ws_comp.cell(row=row_num, column=13, value=f"=C{row_num}").alignment = align_center
        ws_comp.cell(row=row_num, column=14, value=f"=(L{row_num}-M{row_num})^2").alignment = align_right
        
        # WP
        ws_comp.cell(row=row_num, column=15, value=f"=E{row_num}").alignment = align_center
        ws_comp.cell(row=row_num, column=16, value=f"=(L{row_num}-O{row_num})^2").alignment = align_right
        
        # TOPSIS
        ws_comp.cell(row=row_num, column=17, value=f"=G{row_num}").alignment = align_center
        ws_comp.cell(row=row_num, column=18, value=f"=(L{row_num}-Q{row_num})^2").alignment = align_right
        
        is_even = i % 2 == 1
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)
        
        for c in range(11, 19):
            cell = ws_comp.cell(row=row_num, column=c)
            cell.font = font_data
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = box_border
            
    # Add Sum Row
    tot_row = 26
    ws_comp.row_dimensions[tot_row].height = 20
    ws_comp.cell(row=tot_row, column=11, value="TOTAL Σ").font = font_header
    ws_comp.cell(row=tot_row, column=11).fill = fill_header
    ws_comp.cell(row=tot_row, column=11).border = box_border
    ws_comp.cell(row=tot_row, column=11).alignment = align_center
    
    ws_comp.cell(row=tot_row, column=12, value="").border = box_border
    ws_comp.cell(row=tot_row, column=13, value="").border = box_border
    ws_comp.cell(row=tot_row, column=14, value="=SUM(N6:N25)").font = font_bold_data
    ws_comp.cell(row=tot_row, column=14).border = box_border
    ws_comp.cell(row=tot_row, column=14).alignment = align_right
    
    ws_comp.cell(row=tot_row, column=15, value="").border = box_border
    ws_comp.cell(row=tot_row, column=16, value="=SUM(P6:P25)").font = font_bold_data
    ws_comp.cell(row=tot_row, column=16).border = box_border
    ws_comp.cell(row=tot_row, column=16).alignment = align_right
    
    ws_comp.cell(row=tot_row, column=17, value="").border = box_border
    ws_comp.cell(row=tot_row, column=18, value="=SUM(R6:R25)").font = font_bold_data
    ws_comp.cell(row=tot_row, column=18).border = box_border
    ws_comp.cell(row=tot_row, column=18).alignment = align_right
    
    # Table 3: Spearman Result
    ws_comp["K29"] = "Tabel 3. Koefisien Korelasi Spearman (rs)"
    ws_comp["K29"].font = font_section
    ws_comp["K29"].alignment = align_left
    
    rs_headers = ["Metode Pembanding", "Σ d²", "Koefisien rs", "Tingkat Hubungan"]
    ws_comp.row_dimensions[30].height = 24
    for col_idx, h_text in enumerate(rs_headers, 11):
        cell = ws_comp.cell(row=30, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border
        
    metode_pembanding = [
        ("SAW", "N26"),
        ("WP", "P26"),
        ("TOPSIS", "R26")
    ]
    
    for idx, (m_name, sum_cell) in enumerate(metode_pembanding):
        r_num = 31 + idx
        ws_comp.row_dimensions[r_num].height = 20
        ws_comp.cell(row=r_num, column=11, value=f"SWARA-MAIRCA vs {m_name}").font = font_bold_data
        ws_comp.cell(row=r_num, column=11).border = box_border
        ws_comp.cell(row=r_num, column=11).alignment = align_left
        
        ws_comp.cell(row=r_num, column=12, value=f"={sum_cell}").font = font_data
        ws_comp.cell(row=r_num, column=12).border = box_border
        ws_comp.cell(row=r_num, column=12).alignment = align_right
        
        ws_comp.cell(row=r_num, column=13, value=f"=1-(6*L{r_num})/7980").font = font_bold_data
        ws_comp.cell(row=r_num, column=13).number_format = '0.000000'
        ws_comp.cell(row=r_num, column=13).border = box_border
        ws_comp.cell(row=r_num, column=13).alignment = align_right
        
        ws_comp.cell(row=r_num, column=14, value=f'=IF(M{r_num}>0.8,"Sangat Kuat",IF(M{r_num}>0.6,"Kuat",IF(M{r_num}>0.4,"Cukup Kuat","Lemah")))').font = font_bold_data
        ws_comp.cell(row=r_num, column=14).border = box_border
        ws_comp.cell(row=r_num, column=14).alignment = align_center
        
    # Table 4: Sensitivity Table
    ws_comp["A29"] = "Tabel 4. Analisis Sensitivitas (Variasi Bobot KIP - K3)"
    ws_comp["A29"].font = font_section
    ws_comp["A29"].alignment = align_left
    
    sens_headers = ["Variasi K3", "Bobot Baru K3", "Rank 1 Alt.", "Rank Changes", "% Changes"]
    ws_comp.row_dimensions[30].height = 24
    for col_idx, h_text in enumerate(sens_headers, 1):
        cell = ws_comp.cell(row=30, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border
        
    sens_data = [
        ("-20%", 0.1164, "A03", 2, "10.0%"),
        ("-10%", 0.1310, "A03", 2, "10.0%"),
        ("Base (0%)", 0.1456, "A03", 0, "0.0%"),
        ("+10%", 0.1601, "A03", 2, "10.0%"),
        ("+20%", 0.1747, "A03", 2, "10.0%")
    ]
    
    for idx, (var, w_new, alt_1, changes, pct) in enumerate(sens_data):
        r_num = 31 + idx
        ws_comp.row_dimensions[r_num].height = 20
        ws_comp.cell(row=r_num, column=1, value=var).alignment = align_center
        ws_comp.cell(row=r_num, column=2, value=w_new).number_format = '0.0000'
        ws_comp.cell(row=r_num, column=2).alignment = align_right
        ws_comp.cell(row=r_num, column=3, value=alt_1).alignment = align_center
        ws_comp.cell(row=r_num, column=4, value=changes).alignment = align_right
        ws_comp.cell(row=r_num, column=5, value=pct).alignment = align_right
        
        is_even = idx % 2 == 1
        row_fill = fill_highlight if var=="Base (0%)" else (fill_zebra if is_even else PatternFill(fill_type=None))
        
        for c in range(1, 6):
            cell = ws_comp.cell(row=r_num, column=c)
            cell.font = font_bold_data if var=="Base (0%)" else font_data
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = box_border
            
    # Adjust Column Widths for '8.Perbandingan'
    for col in ws_comp.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if val.startswith('='):
                val = " 0.999999 "
            max_len = max(max_len, len(val))
        ws_comp.column_dimensions[col_letter].width = max(max_len + 4, 12)
    # Special columns
    ws_comp.column_dimensions['A'].width = 14
    ws_comp.column_dimensions['B'].width = 12
    ws_comp.column_dimensions['C'].width = 12
    ws_comp.column_dimensions['D'].width = 12
    ws_comp.column_dimensions['E'].width = 12
    ws_comp.column_dimensions['F'].width = 12
    ws_comp.column_dimensions['G'].width = 12
    ws_comp.column_dimensions['H'].width = 14
    ws_comp.column_dimensions['I'].width = 14
    ws_comp.column_dimensions['K'].width = 24
    ws_comp.column_dimensions['L'].width = 14
    ws_comp.column_dimensions['N'].width = 14
    ws_comp.column_dimensions['P'].width = 14
    ws_comp.column_dimensions['R'].width = 14

    wb.save("referensi/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
    print("Successfully updated SWARA_MAIRCA_DETAIL_RUMUS.xlsx programmatically!")

if __name__ == "__main__":
    update_lazizmu_excel()
