import argparse
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_premium_excel(filename="financial_dashboard.xlsx"):
    """Generates a premium corporate financial spreadsheet that looks hand-crafted by a professional analyst."""
    try:
        wb = Workbook()
        
        # 1. Dashboard Tab
        ws = wb.active
        ws.title = "Executive Summary"
        ws.views.sheetView[0].showGridLines = True
        
        # Typography & Color Palette (Classic Corporate Navy & Slate)
        font_family = "Segoe UI"
        
        # Fonts
        font_title = Font(name=font_family, size=18, bold=True, color="1B365D")
        font_subtitle = Font(name=font_family, size=10, italic=True, color="555555")
        font_section = Font(name=font_family, size=12, bold=True, color="1B365D")
        font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        font_kpi_label = Font(name=font_family, size=9, bold=True, color="7F8C8D")
        font_kpi_val = Font(name=font_family, size=16, bold=True, color="1B365D")
        font_data = Font(name=font_family, size=10, color="2C3E50")
        font_total = Font(name=font_family, size=10, bold=True, color="1B365D")
        
        # Fills
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        fill_kpi = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
        fill_total = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
        
        # Borders
        border_light = Side(style='thin', color='BDC3C7')
        border_dark = Side(style='thin', color='2C3E50')
        border_double = Side(style='double', color='1B365D')
        
        box_border = Border(left=border_light, right=border_light, top=border_light, bottom=border_light)
        total_border = Border(top=border_dark, bottom=border_double)
        
        # Alignments
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        # Header / Title Block
        ws.row_dimensions[2].height = 25
        ws["B2"] = "Q1 FINANCIAL PERFORMANCE DASHBOARD"
        ws["B2"].font = font_title
        ws["B2"].alignment = align_left
        
        ws.row_dimensions[3].height = 18
        ws["B3"] = "Published: June 2026 | Prepared by Corporate Finance Division"
        ws["B3"].font = font_subtitle
        ws["B3"].alignment = align_left
        
        # --- KPI Cards Section (Rows 5-7) ---
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 28
        
        # Card 1: Total Revenue
        ws.merge_cells("B5:C5")
        ws.merge_cells("B6:C6")
        ws["B5"] = "TOTAL REVENUE"
        ws["B6"] = "=SUM(E11:E15)"
        
        # Card 2: Units Sold
        ws.merge_cells("E5:F5")
        ws.merge_cells("E6:F6")
        ws["E5"] = "UNITS DELIVERED"
        ws["E6"] = "=SUM(C11:C15)"
        
        # Card 3: Average Unit Price
        ws.merge_cells("H5:I5")
        ws.merge_cells("H6:I6")
        ws["H5"] = "AVG CONTRACT VALUE"
        ws["H6"] = "=AVERAGE(D11:D15)"
        
        # Style KPI Cards
        kpi_ranges = [("B5", "B6", 2, 3), ("E5", "E6", 5, 6), ("H5", "H6", 8, 9)]
        for start_lbl, start_val, col_start, col_end in kpi_ranges:
            ws[start_lbl].font = font_kpi_label
            ws[start_lbl].alignment = align_center
            ws[start_val].font = font_kpi_val
            ws[start_val].alignment = align_center
            ws[start_val].number_format = '$#,##0.00' if start_lbl != "TOTAL REVENUE" else '$#,##0.00'
            if start_lbl == "UNITS DELIVERED":
                ws[start_val].number_format = '#,##0'
                
            for r in range(5, 7):
                for c in range(col_start, col_end + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = fill_kpi
                    cell.border = box_border

        # --- Section Title (Row 9) ---
        ws.row_dimensions[9].height = 22
        ws["B9"] = "Detailed Performance by Product Line"
        ws["B9"].font = font_section
        ws["B9"].alignment = align_left
        
        # --- Data Table (Rows 10-16) ---
        headers = ["Product / Service", "Category", "Units Sold", "Unit Price", "Total Revenue"]
        ws.row_dimensions[10].height = 25
        
        for idx, h_text in enumerate(headers, 2):  # Start at Column B (2)
            cell = ws.cell(row=10, column=idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if idx > 3 else align_left
            cell.border = box_border
            
        data = [
            ["Cloud Infrastructure (SaaS)", "Subscription", 1450, 120.00],
            ["Cybersecurity Suite Pro", "License", 320, 850.00],
            ["Data Analytics Platform", "License", 185, 2400.00],
            ["Enterprise Custom Dev", "Services", 12, 15000.00],
            ["Dedicated Technical Support", "Services", 48, 1250.00]
        ]
        
        for r_offset, row_data in enumerate(data):
            r_idx = 11 + r_offset
            ws.row_dimensions[r_idx].height = 20
            
            # Product, Category
            ws.cell(row=r_idx, column=2, value=row_data[0]).alignment = align_left
            ws.cell(row=r_idx, column=3, value=row_data[1]).alignment = align_left
            
            # Units
            units_cell = ws.cell(row=r_idx, column=4, value=row_data[2])
            units_cell.alignment = align_right
            units_cell.number_format = '#,##0'
            
            # Unit Price
            price_cell = ws.cell(row=r_idx, column=5, value=row_data[3])
            price_cell.alignment = align_right
            price_cell.number_format = '$#,##0.00'
            
            # Total Revenue (Formula)
            rev_cell = ws.cell(row=r_idx, column=6, value=f"=D{r_idx}*E{r_idx}")
            rev_cell.alignment = align_right
            rev_cell.number_format = '$#,##0.00'
            
            # Styling & Zebra Striping
            for col_idx in range(2, 7):
                c = ws.cell(row=r_idx, column=col_idx)
                c.font = font_data
                c.border = box_border
                if r_offset % 2 == 1:
                    c.fill = fill_zebra
                    
        # Total Summary Row (Row 16)
        t_row = 16
        ws.row_dimensions[t_row].height = 24
        
        ws.cell(row=t_row, column=2, value="Total Summary").alignment = align_left
        ws.cell(row=t_row, column=3, value="").alignment = align_left
        
        ws.cell(row=t_row, column=4, value=f"=SUM(D11:D{t_row-1})").alignment = align_right
        ws.cell(row=t_row, column=4).number_format = '#,##0'
        
        ws.cell(row=t_row, column=5, value=f"=AVERAGE(E11:E{t_row-1})").alignment = align_right
        ws.cell(row=t_row, column=5).number_format = '$#,##0.00'
        
        ws.cell(row=t_row, column=6, value=f"=SUM(F11:F{t_row-1})").alignment = align_right
        ws.cell(row=t_row, column=6).number_format = '$#,##0.00'
        
        for col_idx in range(2, 7):
            c = ws.cell(row=t_row, column=col_idx)
            c.font = font_total
            c.fill = fill_total
            c.border = total_border
            
        # Adjust Column Widths dynamically
        for col in ws.columns:
            # Skip Column A
            if col[0].column == 1:
                ws.column_dimensions[get_column_letter(1)].width = 3
                continue
                
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if val.startswith('='):
                    val = " $9,999,999.00 "
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        wb.save(filename)
        print(f"Premium Excel workbook successfully created: {filename}")
    except Exception as e:
        print(f"Error creating premium Excel: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Create a premium formatted Excel spreadsheet.")
    parser.add_argument("-o", "--output", default="financial_dashboard.xlsx", help="Filename of the spreadsheet to generate.")
    args = parser.parse_args()
    create_premium_excel(args.output)
