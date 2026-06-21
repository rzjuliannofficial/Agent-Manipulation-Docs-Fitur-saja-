import openpyxl
import sys
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def calculate_saw_wp_topsis(alt_data, c_types, weights):
    num_alt = len(alt_data)
    num_crit = len(c_types)
    
    # --- SAW ---
    max_vals = [max(alt_data[r][c] for r in range(num_alt)) for c in range(num_crit)]
    min_vals = [min(alt_data[r][c] for r in range(num_alt)) for c in range(num_crit)]
    
    saw_v = []
    for r in range(num_alt):
        val = 0.0
        for c in range(num_crit):
            is_benefit = c_types[c].strip().lower() == "benefit"
            if is_benefit:
                norm = alt_data[r][c] / max_vals[c] if max_vals[c] != 0 else 0.0
            else:
                norm = min_vals[c] / alt_data[r][c] if alt_data[r][c] != 0 else 0.0
            val += norm * weights[c]
        saw_v.append(val)
        
    # --- WP ---
    wp_v = []
    for r in range(num_alt):
        val = 1.0
        for c in range(num_crit):
            is_benefit = c_types[c].strip().lower() == "benefit"
            power = weights[c] if is_benefit else -weights[c]
            # Avoid division by zero by using a tiny offset if val is 0
            base_val = alt_data[r][c] if alt_data[r][c] != 0 else 1e-9
            val *= (base_val ** power)
        wp_v.append(val)
        
    # --- TOPSIS ---
    dividers = []
    for c in range(num_crit):
        sum_sq = sum(alt_data[r][c]**2 for r in range(num_alt))
        dividers.append(sum_sq**0.5)
        
    v_matrix = []
    for r in range(num_alt):
        row_v = []
        for c in range(num_crit):
            norm = alt_data[r][c] / dividers[c] if dividers[c] != 0 else 0.0
            row_v.append(norm * weights[c])
        v_matrix.append(row_v)
        
    a_plus = []
    a_minus = []
    for c in range(num_crit):
        col_v = [v_matrix[r][c] for r in range(num_alt)]
        is_benefit = c_types[c].strip().lower() == "benefit"
        if is_benefit:
            a_plus.append(max(col_v))
            a_minus.append(min(col_v))
        else:
            a_plus.append(min(col_v))
            a_minus.append(max(col_v))
            
    topsis_v = []
    for r in range(num_alt):
        d_plus = sum((v_matrix[r][c] - a_plus[c])**2 for c in range(num_crit))**0.5
        d_minus = sum((v_matrix[r][c] - a_minus[c])**2 for c in range(num_crit))**0.5
        if (d_plus + d_minus) == 0:
            val = 0.0
        else:
            val = d_minus / (d_plus + d_minus)
        topsis_v.append(val)
        
    return saw_v, wp_v, topsis_v

def run_mcdm_engine(input_path, output_path):
    try:
        wb = openpyxl.load_workbook(input_path)
        
        # 1. Read Criteria Data
        ws_data = wb["1.Data"]
        ws_data.views.sheetView[0].showGridLines = True
        
        # Style definition for professional look (Segoe UI Corporate Theme)
        font_family = "Segoe UI"
        font_title = Font(name=font_family, size=14, bold=True, color="1E293B")
        font_subtitle = Font(name=font_family, size=10, italic=True, color="64748B")
        font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        font_data = Font(name=font_family, size=10, color="334155")
        font_bold = Font(name=font_family, size=10, bold=True, color="1E293B")
        
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_highlight_green = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # soft mint
        font_highlight_green = Font(name=font_family, size=10, bold=True, color="065F46")
        fill_base_variation = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # soft slate
        
        border_horizontal_light = Side(style='thin', color='E2E8F0')
        border_double_bottom = Side(style='double', color='1E293B')
        border_top_thin = Side(style='thin', color='1E293B')
        
        box_border = Border(bottom=border_horizontal_light) # horizontal lines only
        header_border = Border(top=border_top_thin, bottom=Side(style='medium', color='1E293B'))
        summary_border = Border(top=border_top_thin, bottom=border_double_bottom)
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # Read headers (Definitions of criteria start at column 4 in row 2-4)
        c_codes = []
        c_types = []
        c_names = []
        col = 4
        while True:
            val = ws_data.cell(row=2, column=col).value
            if not val:
                break
            c_codes.append(val)
            c_types.append(ws_data.cell(row=3, column=col).value)
            c_names.append(ws_data.cell(row=4, column=col).value)
            col += 1
            
        num_criteria = len(c_codes)
        
        # Read Alternatives (A01-A20)
        alt_codes = []
        alt_data = []
        row = 9
        while True:
            val = ws_data.cell(row=row, column=1).value
            if not val:
                break
            alt_codes.append(val)
            # Read criteria values from column 2 (K1) to column 2 + num_criteria - 1
            row_vals = []
            for c_idx in range(num_criteria):
                row_vals.append(float(ws_data.cell(row=row, column=c_idx + 2).value or 0.0))
            alt_data.append(row_vals)
            row += 1
            
        num_alternatives = len(alt_codes)
        last_alt_row = 9 + num_alternatives - 1
        
        # Calculate Max/Min values per criterion for normalisation mapping
        max_vals = []
        min_vals = []
        for c_idx in range(num_criteria):
            col_vals = [alt_data[r_idx][c_idx] for r_idx in range(num_alternatives)]
            max_vals.append(max(col_vals))
            min_vals.append(min(col_vals))
            
        # Write MAX/MIN dynamic formulas back to 1.Data
        max_row = last_alt_row + 2
        min_row = last_alt_row + 3
        
        ws_data.cell(row=max_row, column=1, value="Nilai MAX").font = font_bold
        ws_data.cell(row=max_row, column=1).alignment = align_left
        ws_data.cell(row=max_row, column=1).border = summary_border
        
        ws_data.cell(row=min_row, column=1, value="Nilai MIN").font = font_bold
        ws_data.cell(row=min_row, column=1).alignment = align_left
        ws_data.cell(row=min_row, column=1).border = summary_border
        
        for c_idx in range(num_criteria):
            col_let = get_column_letter(c_idx + 2)
            
            cell_max = ws_data.cell(row=max_row, column=c_idx + 2, value=f"=MAX({col_let}9:{col_let}{last_alt_row})")
            cell_max.font = font_bold
            cell_max.alignment = align_right
            cell_max.border = summary_border
            
            cell_min = ws_data.cell(row=min_row, column=c_idx + 2, value=f"=MIN({col_let}9:{col_let}{last_alt_row})")
            cell_min.font = font_bold
            cell_min.alignment = align_right
            cell_min.border = summary_border
            
        # Format 1.Data headers and borders
        for r_idx in range(9, last_alt_row + 1):
            ws_data.row_dimensions[r_idx].height = 22
            ws_data.cell(row=r_idx, column=1).font = font_bold
            ws_data.cell(row=r_idx, column=1).alignment = align_center
            ws_data.cell(row=r_idx, column=1).border = box_border
            for c_idx in range(num_criteria):
                cell = ws_data.cell(row=r_idx, column=c_idx + 2)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = box_border
                    
        # Apply style to headers of 1.Data
        for c_idx in range(num_criteria + 1):
            for r_idx in range(2, 5):
                cell = ws_data.cell(row=r_idx, column=c_idx + 1)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = header_border
            
        # 2. SWARA Sheet Weights calculation with dynamic formula
        ws_swara = wb["2.SWARA"]
        ws_swara.views.sheetView[0].showGridLines = True
        
        # Populate headers in SWARA
        ws_swara.cell(row=3, column=5, value="Kj = Sj + 1")
        ws_swara.cell(row=3, column=6, value="Qj")
        ws_swara.cell(row=3, column=7, value="Wj (Bobot)")
        
        for c_idx in range(1, 8):
            cell = ws_swara.cell(row=3, column=c_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
        
        # Read Sj inputs from sheet for local python calculation
        swara_rows = []
        for r_idx in range(4, 4 + num_criteria):
            prio = ws_swara.cell(row=r_idx, column=1).value
            code = ws_swara.cell(row=r_idx, column=2).value
            name = ws_swara.cell(row=r_idx, column=3).value
            sj = float(ws_swara.cell(row=r_idx, column=4).value or 0.0)
            swara_rows.append({"prio": prio, "code": code, "name": name, "sj": sj, "row_idx": r_idx})
            
        # Local weights calculation to use in SAW/WP/TOPSIS/MAIRCA python engines
        swara_rows[0]["kj"] = 1.0
        swara_rows[0]["qj"] = 1.0
        for i in range(1, len(swara_rows)):
            swara_rows[i]["kj"] = swara_rows[i]["sj"] + 1.0
            swara_rows[i]["qj"] = swara_rows[i - 1]["qj"] / swara_rows[i]["kj"]
            
        sum_qj = sum(item["qj"] for item in swara_rows)
        for item in swara_rows:
            item["wj"] = item["qj"] / sum_qj
            
        # Map weights back to c_codes
        weights_map = {item["code"]: item["wj"] for item in swara_rows}
        ordered_weights = [weights_map[code] for code in c_codes]
        
        # Write SWARA calculation formulas
        for r_idx in range(4, 4 + num_criteria):
            ws_swara.row_dimensions[r_idx].height = 22
            for c_idx in range(1, 8):
                cell = ws_swara.cell(row=r_idx, column=c_idx)
                cell.font = font_data
                cell.border = box_border
            
            ws_swara.cell(row=r_idx, column=1).alignment = align_center
            ws_swara.cell(row=r_idx, column=2).alignment = align_center
            ws_swara.cell(row=r_idx, column=3).alignment = align_left
            ws_swara.cell(row=r_idx, column=4).alignment = align_right
            ws_swara.cell(row=r_idx, column=4).number_format = '0.00'
            
            if r_idx == 4:
                ws_swara.cell(row=r_idx, column=5, value=1.0)
                ws_swara.cell(row=r_idx, column=6, value=1.0)
            else:
                ws_swara.cell(row=r_idx, column=5, value=f"=D{r_idx}+1")
                ws_swara.cell(row=r_idx, column=6, value=f"=F{r_idx-1}/E{r_idx}")
            
            ws_swara.cell(row=r_idx, column=7, value=f"=F{r_idx}/SUM(F$4:F${4+num_criteria-1})")
            
            ws_swara.cell(row=r_idx, column=5).alignment = align_right
            ws_swara.cell(row=r_idx, column=5).number_format = '0.00'
            ws_swara.cell(row=r_idx, column=6).alignment = align_right
            ws_swara.cell(row=r_idx, column=6).number_format = '0.0000'
            ws_swara.cell(row=r_idx, column=7).alignment = align_right
            ws_swara.cell(row=r_idx, column=7).number_format = '0.0000'
                    
        # Rekap Bobot SWARA (Urutan K1 ke K_n)
        rekap_start = 4 + num_criteria + 4
        ws_swara.cell(row=rekap_start, column=1, value="REKAP BOBOT SWARA — Urutan K1 s/d K_n").font = font_title
        
        ws_swara.cell(row=rekap_start+1, column=1, value="No")
        ws_swara.cell(row=rekap_start+1, column=2, value="Kode")
        ws_swara.cell(row=rekap_start+1, column=3, value="Nama Kriteria")
        ws_swara.cell(row=rekap_start+1, column=4, value="Tipe")
        ws_swara.cell(row=rekap_start+1, column=5, value="Wj (Bobot)")
        
        for c_idx in range(1, 6):
            cell = ws_swara.cell(row=rekap_start+1, column=c_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        for i in range(num_criteria):
            curr_row = rekap_start + 2 + i
            ws_swara.row_dimensions[curr_row].height = 22
            ws_swara.cell(row=curr_row, column=1, value=i+1).alignment = align_center
            ws_swara.cell(row=curr_row, column=2, value=c_codes[i]).alignment = align_center
            ws_swara.cell(row=curr_row, column=3, value=c_names[i]).alignment = align_left
            ws_swara.cell(row=curr_row, column=4, value=c_types[i]).alignment = align_center
            
            # Lookup formula to fetch weight dynamically from the SWARA calculation table
            cell_w = ws_swara.cell(row=curr_row, column=5, value=f"=VLOOKUP(B{curr_row}, B$4:G${4+num_criteria-1}, 6, FALSE)")
            cell_w.alignment = align_right
            cell_w.number_format = '0.0000'
            
            for c_idx in range(1, 6):
                cell = ws_swara.cell(row=curr_row, column=c_idx)
                cell.font = font_data
                cell.border = box_border
                    
        # 3. Normalisasi (MAIRCA) - Dynamic Formula
        if "3.Normalisasi" in wb.sheetnames:
            del wb["3.Normalisasi"]
        ws_norm = wb.create_sheet("3.Normalisasi")
        ws_norm.views.sheetView[0].showGridLines = True
        
        ws_norm["A1"] = "NORMALISASI MATRIKS KEPUTUSAN — MAIRCA"
        ws_norm["A1"].font = font_title
        ws_norm["A2"] = "Benefit: (xij-min)/(max-min) | Cost: (max-xij)/(max-min)"
        ws_norm["A2"].font = font_subtitle
        
        ws_norm["A4"] = "Alternatif"
        ws_norm["A4"].font = font_header
        ws_norm["A4"].fill = fill_header
        ws_norm["A4"].alignment = align_center
        ws_norm["A4"].border = header_border
        
        for i, code in enumerate(c_codes):
            cell = ws_norm.cell(row=4, column=i+2, value=code)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        norm_data = []
        for r_idx, alt_code in enumerate(alt_codes):
            ws_norm.row_dimensions[5 + r_idx].height = 22
            ws_norm.cell(row=5 + r_idx, column=1, value=alt_code).font = font_bold
            ws_norm.cell(row=5 + r_idx, column=1).alignment = align_center
            ws_norm.cell(row=5 + r_idx, column=1).border = box_border
            
            row_norm = []
            for c_idx, code in enumerate(c_codes):
                col_let = get_column_letter(c_idx + 2)
                is_benefit = c_types[c_idx].strip().lower() == "benefit"
                data_row = 9 + r_idx
                
                # Local calculation for python flow
                x_val = alt_data[r_idx][c_idx]
                mn = min_vals[c_idx]
                mx = max_vals[c_idx]
                n_val = (x_val - mn) / (mx - mn) if mx != mn else 1.0
                if not is_benefit:
                    n_val = 1.0 - n_val
                row_norm.append(n_val)
                
                # Excel formula
                if is_benefit:
                    formula = f"=( '1.Data'!{col_let}{data_row} - '1.Data'!{col_let}${min_row} ) / ( '1.Data'!{col_let}${max_row} - '1.Data'!{col_let}${min_row} )"
                else:
                    formula = f"=( '1.Data'!{col_let}${max_row} - '1.Data'!{col_let}{data_row} ) / ( '1.Data'!{col_let}${max_row} - '1.Data'!{col_let}${min_row} )"
                
                cell = ws_norm.cell(row=5 + r_idx, column=c_idx + 2, value=formula)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = box_border
                cell.number_format = '0.0000'
            norm_data.append(row_norm)
                    
        # 4. Preferensi Teoritis - TPij = Wj / m
        if "4.Pref.Teoritis" in wb.sheetnames:
            del wb["4.Pref.Teoritis"]
        ws_tp = wb.create_sheet("4.Pref.Teoritis")
        ws_tp.views.sheetView[0].showGridLines = True
        
        ws_tp["A1"] = "PREFERENSI TEORITIS (TPij) — MAIRCA"
        ws_tp["A1"].font = font_title
        ws_tp["A2"] = f"Rumus: TPij = Wj / m  (m = {num_alternatives})"
        ws_tp["A2"].font = font_subtitle
        
        ws_tp["A4"] = "Alternatif"
        ws_tp["A4"].font = font_header
        ws_tp["A4"].fill = fill_header
        ws_tp["A4"].alignment = align_center
        ws_tp["A4"].border = header_border
        
        for i, code in enumerate(c_codes):
            cell = ws_tp.cell(row=4, column=i+2, value=code)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        tp_row = [wj / num_alternatives for wj in ordered_weights]
        for r_idx, alt_code in enumerate(alt_codes):
            ws_tp.row_dimensions[5 + r_idx].height = 22
            ws_tp.cell(row=5 + r_idx, column=1, value=alt_code).font = font_bold
            ws_tp.cell(row=5 + r_idx, column=1).alignment = align_center
            ws_tp.cell(row=5 + r_idx, column=1).border = box_border
            
            for c_idx in range(num_criteria):
                rekap_row = rekap_start + 2 + c_idx
                formula = f"='2.SWARA'!E{rekap_row} / {num_alternatives}"
                
                cell = ws_tp.cell(row=5 + r_idx, column=c_idx + 2, value=formula)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = box_border
                cell.number_format = '0.0000'
                    
        # 5. Preferensi Nyata - PPij = TPij * Normalisasi
        if "5.Pref.Nyata" in wb.sheetnames:
            del wb["5.Pref.Nyata"]
        ws_pp = wb.create_sheet("5.Pref.Nyata")
        ws_pp.views.sheetView[0].showGridLines = True
        
        ws_pp["A1"] = "PREFERENSI NYATA (PPij) — MAIRCA"
        ws_pp["A1"].font = font_title
        ws_pp["A2"] = "Rumus: PPij = TPij * x_ij_norm"
        ws_pp["A2"].font = font_subtitle
        
        ws_pp["A4"] = "Alternatif"
        ws_pp["A4"].font = font_header
        ws_pp["A4"].fill = fill_header
        ws_pp["A4"].alignment = align_center
        ws_pp["A4"].border = header_border
        
        for i, code in enumerate(c_codes):
            cell = ws_pp.cell(row=4, column=i+2, value=code)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        pp_data = []
        for r_idx, alt_code in enumerate(alt_codes):
            ws_pp.row_dimensions[5 + r_idx].height = 22
            ws_pp.cell(row=5 + r_idx, column=1, value=alt_code).font = font_bold
            ws_pp.cell(row=5 + r_idx, column=1).alignment = align_center
            ws_pp.cell(row=5 + r_idx, column=1).border = box_border
            
            row_pp = []
            for c_idx in range(num_criteria):
                col_let = get_column_letter(c_idx + 2)
                row_num = 5 + r_idx
                formula = f"='4.Pref.Teoritis'!{col_let}{row_num} * '3.Normalisasi'!{col_let}{row_num}"
                
                cell = ws_pp.cell(row=5 + r_idx, column=c_idx + 2, value=formula)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = box_border
                cell.number_format = '0.0000'
                row_pp.append(tp_row[c_idx] * row_norm[c_idx])
            pp_data.append(row_pp)
                    
        # 6. Gap & Qi - Gij = TPij - PPij | Qi = sum(Gij)
        if "6.Gap" in wb.sheetnames:
            del wb["6.Gap"]
        ws_gap = wb.create_sheet("6.Gap")
        ws_gap.views.sheetView[0].showGridLines = True
        
        ws_gap["A1"] = "MATRIKS GAP (Gij) DAN NILAI Qi — MAIRCA"
        ws_gap["A1"].font = font_title
        ws_gap["A2"] = "Rumus: Gij = TPij - PPij | Qi = sum(Gij)"
        ws_gap["A2"].font = font_subtitle
        
        ws_gap["A4"] = "Alternatif"
        ws_gap["A4"].font = font_header
        ws_gap["A4"].fill = fill_header
        ws_gap["A4"].alignment = align_center
        ws_gap["A4"].border = header_border
        
        for i, code in enumerate(c_codes):
            cell = ws_gap.cell(row=4, column=i+2, value=code)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        col_qi = num_criteria + 2
        col_rank = num_criteria + 3
        
        cell_qi_hdr = ws_gap.cell(row=4, column=col_qi, value="Qi (Σ Gap)")
        cell_qi_hdr.font = font_header
        cell_qi_hdr.fill = fill_header
        cell_qi_hdr.alignment = align_center
        cell_qi_hdr.border = header_border
        
        cell_rnk_hdr = ws_gap.cell(row=4, column=col_rank, value="Rank")
        cell_rnk_hdr.font = font_header
        cell_rnk_hdr.fill = fill_header
        cell_rnk_hdr.alignment = align_center
        cell_rnk_hdr.border = header_border
        
        qi_list = []
        for r_idx, alt_code in enumerate(alt_codes):
            ws_gap.row_dimensions[5 + r_idx].height = 22
            ws_gap.cell(row=5 + r_idx, column=1, value=alt_code).font = font_bold
            ws_gap.cell(row=5 + r_idx, column=1).alignment = align_center
            ws_gap.cell(row=5 + r_idx, column=1).border = box_border
            
            row_num = 5 + r_idx
            row_gap_sum = 0.0
            for c_idx in range(num_criteria):
                col_let = get_column_letter(c_idx + 2)
                # Local gap
                g_val = tp_row[c_idx] - (tp_row[c_idx] * norm_data[r_idx][c_idx])
                row_gap_sum += g_val
                
                # Excel formula
                formula = f"='4.Pref.Teoritis'!{col_let}{row_num} - '5.Pref.Nyata'!{col_let}{row_num}"
                
                cell = ws_gap.cell(row=5 + r_idx, column=c_idx + 2, value=formula)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = box_border
                cell.number_format = '0.0000'
                    
            # Qi = Sum of gaps
            last_col_let = get_column_letter(num_criteria + 1)
            cell_qi = ws_gap.cell(row=row_num, column=col_qi, value=f"=SUM(B{row_num}:{last_col_let}{row_num})")
            cell_qi.font = font_bold
            cell_qi.alignment = align_right
            cell_qi.border = box_border
            cell_qi.number_format = '0.0000'
            
            qi_list.append({"alt": alt_code, "qi": row_gap_sum, "row": row_num})
            
            # Rank = RANK(Qi, Qi_Range, 1)
            qi_col_let = get_column_letter(col_qi)
            cell_rank = ws_gap.cell(row=row_num, column=col_rank, value=f"=RANK({qi_col_let}{row_num}, {qi_col_let}$5:{qi_col_let}${5 + num_alternatives - 1}, 1)")
            cell_rank.font = font_bold
            cell_rank.alignment = align_center
            cell_rank.border = box_border
            
        # Calculate local ranks for rank 1 sensitivity comparison in python
        sorted_qi = sorted(qi_list, key=lambda x: x["qi"])
        for rank, item in enumerate(sorted_qi, start=1):
            item["rank"] = rank
        base_ranks = {item["alt"]: item["rank"] for item in qi_list}
        
        # 7. Hasil - Final Rank summary dashboard
        if "7.Hasil" in wb.sheetnames:
            del wb["7.Hasil"]
        ws_hasil = wb.create_sheet("7.Hasil")
        ws_hasil.views.sheetView[0].showGridLines = True
        
        ws_hasil["A1"] = "HASIL AKHIR PERANGKINGAN — METODE SWARA + MAIRCA"
        ws_hasil["A1"].font = font_title
        ws_hasil["A2"] = "Catatan: Nilai Qi terkecil = alternatif terbaik (Peringkat 1)"
        ws_hasil["A2"].font = font_subtitle
        
        headers_hasil = ["No", "Alternatif", "Nilai Qi", "Rank", "Keterangan"]
        ws_hasil.row_dimensions[4].height = 28
        for idx, h_text in enumerate(headers_hasil, 1):
            cell = ws_hasil.cell(row=4, column=idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        qi_col_let = get_column_letter(col_qi)
        rank_col_let = get_column_letter(col_rank)
        
        for idx in range(1, num_alternatives + 1):
            row_num = 4 + idx
            ws_hasil.row_dimensions[row_num].height = 22
            
            ws_hasil.cell(row=row_num, column=1, value=idx).alignment = align_center
            
            cell_qi = ws_hasil.cell(row=row_num, column=3, value=f"=SMALL('6.Gap'!${qi_col_let}$5:${qi_col_let}${5 + num_alternatives - 1}, A{row_num})")
            cell_qi.alignment = align_right
            cell_qi.number_format = '0.0000'
            
            cell_alt = ws_hasil.cell(row=row_num, column=2, value=f"=INDEX('6.Gap'!$A$5:$A${5 + num_alternatives - 1}, MATCH(C{row_num}, '6.Gap'!${qi_col_let}$5:${qi_col_let}${5 + num_alternatives - 1}, 0))")
            cell_alt.alignment = align_center
            
            cell_rank = ws_hasil.cell(row=row_num, column=4, value=f"=INDEX('6.Gap'!${rank_col_let}$5:${rank_col_let}${5 + num_alternatives - 1}, MATCH(C{row_num}, '6.Gap'!${qi_col_let}$5:${qi_col_let}${5 + num_alternatives - 1}, 0))")
            cell_rank.alignment = align_center
            
            cell_desc = ws_hasil.cell(row=row_num, column=5, value=f'=IF(D{row_num}=1, "★ Peringkat 1 – Sangat Direkomendasikan", IF(D{row_num}<=5, "Sangat Direkomendasikan", IF(D{row_num}<=10, "Direkomendasikan", "Pertimbangkan")))')
            cell_desc.alignment = align_left
            
            is_rank_1 = (idx == 1)
            for c_idx in range(1, 6):
                c = ws_hasil.cell(row=row_num, column=c_idx)
                if is_rank_1:
                    c.font = font_highlight_green
                    c.fill = fill_highlight_green
                else:
                    c.font = font_data
                c.border = box_border
                
            ws_hasil.cell(row=row_num, column=1).font = font_highlight_green if is_rank_1 else font_bold
            ws_hasil.cell(row=row_num, column=2).font = font_highlight_green if is_rank_1 else font_bold

        # 8. Perbandingan & Sensitivitas (Dynamic computation)
        if "8.Perbandingan" in wb.sheetnames:
            del wb["8.Perbandingan"]
        ws_perb = wb.create_sheet("8.Perbandingan")
        ws_perb.views.sheetView[0].showGridLines = True
        
        ws_perb["A1"] = "ANALISIS PERBANDINGAN METODE DAN UJI SENSITIVITAS"
        ws_perb["A1"].font = font_title
        ws_perb["A2"] = "Berdasarkan Data Seleksi Smartphone Terbaik (20 Alternatif)"
        ws_perb["A2"].font = font_subtitle
        
        ws_perb["A4"] = "Tabel 1. Perbandingan Nilai dan Peringkat SAW, WP, TOPSIS, dan SWARA-MAIRCA"
        ws_perb["A4"].font = font_bold
        
        # Calculate comparison method values via python engine
        saw_v, wp_v, topsis_v = calculate_saw_wp_topsis(alt_data, c_types, ordered_weights)
        
        headers_perb = ["Alternatif", "SAW V", "Rank SAW", "WP V", "Rank WP", "TOPSIS V", "Rank TOPSIS", "MAIRCA Qi", "Rank MAIRCA"]
        ws_perb.row_dimensions[5].height = 28
        for col_idx, h_text in enumerate(headers_perb, 1):
            cell = ws_perb.cell(row=5, column=col_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        for r_idx, alt_code in enumerate(alt_codes):
            row_num = 6 + r_idx
            ws_perb.row_dimensions[row_num].height = 22
            
            ws_perb.cell(row=row_num, column=1, value=alt_code).font = font_bold
            ws_perb.cell(row=row_num, column=1).alignment = align_center
            ws_perb.cell(row=row_num, column=1).border = box_border
            
            # SAW value & rank
            c_saw_v = ws_perb.cell(row=row_num, column=2, value=saw_v[r_idx])
            c_saw_v.number_format = '0.0000'
            c_saw_v.alignment = align_right
            c_saw_r = ws_perb.cell(row=row_num, column=3, value=f"=RANK(B{row_num},$B$6:$B$25,0)")
            c_saw_r.alignment = align_center
            
            # WP value & rank
            c_wp_v = ws_perb.cell(row=row_num, column=4, value=wp_v[r_idx])
            c_wp_v.number_format = '0.0000'
            c_wp_v.alignment = align_right
            c_wp_r = ws_perb.cell(row=row_num, column=5, value=f"=RANK(D{row_num},$D$6:$D$25,0)")
            c_wp_r.alignment = align_center
            
            # TOPSIS value & rank
            c_top_v = ws_perb.cell(row=row_num, column=6, value=topsis_v[r_idx])
            c_top_v.number_format = '0.0000'
            c_top_v.alignment = align_right
            c_top_r = ws_perb.cell(row=row_num, column=7, value=f"=RANK(F{row_num},$F$6:$F$25,0)")
            c_top_r.alignment = align_center
            
            # MAIRCA references
            c_ma_qi = ws_perb.cell(row=row_num, column=8, value=f"='6.Gap'!{qi_col_let}{row_num-1}")
            c_ma_qi.number_format = '0.0000'
            c_ma_qi.alignment = align_right
            c_ma_r = ws_perb.cell(row=row_num, column=9, value=f"='6.Gap'!{rank_col_let}{row_num-1}")
            c_ma_r.alignment = align_center
            
            for c_idx in range(2, 10):
                cell = ws_perb.cell(row=row_num, column=c_idx)
                cell.font = font_data
                cell.border = box_border
                    
        # Spearman Column Setup (cols 11 to 18)
        headers_spearman = ["Alt", "Rank R1\n(MAIRCA)", "Rank R2\n(SAW)", "d² (SAW)", "Rank R3\n(WP)", "d² (WP)", "Rank R4\n(TOPSIS)", "d² (TOPSIS)"]
        for col_offset, h_text in enumerate(headers_spearman):
            cell = ws_perb.cell(row=5, column=11 + col_offset, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        for r_idx, alt_code in enumerate(alt_codes):
            row_num = 6 + r_idx
            # Alt
            ws_perb.cell(row=row_num, column=11, value=alt_code).font = font_bold
            ws_perb.cell(row=row_num, column=11).alignment = align_center
            ws_perb.cell(row=row_num, column=11).border = box_border
            
            # Rank R1 (MAIRCA)
            ws_perb.cell(row=row_num, column=12, value=f"=I{row_num}").alignment = align_center
            # Rank R2 (SAW)
            ws_perb.cell(row=row_num, column=13, value=f"=C{row_num}").alignment = align_center
            # d^2 (SAW)
            c_d_saw = ws_perb.cell(row=row_num, column=14, value=f"=(L{row_num}-M{row_num})^2")
            c_d_saw.alignment = align_right
            c_d_saw.number_format = '0'
            
            # Rank R3 (WP)
            ws_perb.cell(row=row_num, column=15, value=f"=E{row_num}").alignment = align_center
            # d^2 (WP)
            c_d_wp = ws_perb.cell(row=row_num, column=16, value=f"=(L{row_num}-O{row_num})^2")
            c_d_wp.alignment = align_right
            c_d_wp.number_format = '0'
            
            # Rank R4 (TOPSIS)
            ws_perb.cell(row=row_num, column=17, value=f"=G{row_num}").alignment = align_center
            # d^2 (TOPSIS)
            c_d_top = ws_perb.cell(row=row_num, column=18, value=f"=(L{row_num}-Q{row_num})^2")
            c_d_top.alignment = align_right
            c_d_top.number_format = '0'
            
            for c_offset in range(1, 8):
                cell = ws_perb.cell(row=row_num, column=11 + c_offset)
                cell.font = font_data
                cell.border = box_border
                    
        # Total Sum row for Spearman
        ws_perb.cell(row=26, column=11, value="TOTAL Σ").font = font_bold
        ws_perb.cell(row=26, column=11).border = summary_border
        ws_perb.cell(row=26, column=11).alignment = align_center
        
        # SUM(d^2 SAW)
        cell_sum_saw = ws_perb.cell(row=26, column=14, value="=SUM(N6:N25)")
        cell_sum_saw.font = font_bold
        cell_sum_saw.alignment = align_right
        cell_sum_saw.number_format = '0'
        cell_sum_saw.border = summary_border
        
        # SUM(d^2 WP)
        cell_sum_wp = ws_perb.cell(row=26, column=16, value="=SUM(P6:P25)")
        cell_sum_wp.font = font_bold
        cell_sum_wp.alignment = align_right
        cell_sum_wp.number_format = '0'
        cell_sum_wp.border = summary_border
        
        # SUM(d^2 TOPSIS)
        cell_sum_top = ws_perb.cell(row=26, column=18, value="=SUM(R6:R25)")
        cell_sum_top.font = font_bold
        cell_sum_top.alignment = align_right
        cell_sum_top.number_format = '0'
        cell_sum_top.border = summary_border
        
        for col_idx in [12, 13, 15, 17]:
            ws_perb.cell(row=26, column=col_idx).border = summary_border
            
        # Spearman table
        ws_perb.cell(row=29, column=11, value="Tabel 3. Koefisien Korelasi Spearman (rs)").font = font_bold
        
        headers_rs = ["Metode Pembanding", "Σ d²", "Koefisien rs", "Tingkat Hubungan"]
        for col_offset, h_text in enumerate(headers_rs):
            cell = ws_perb.cell(row=30, column=11 + col_offset, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        # SAW comparison row
        ws_perb.cell(row=31, column=11, value="SWARA-MAIRCA vs SAW").alignment = align_left
        ws_perb.cell(row=31, column=12, value="=N26").alignment = align_right
        ws_perb.cell(row=31, column=12).number_format = '0'
        ws_perb.cell(row=31, column=13, value="=1-(6*L31)/7980")
        ws_perb.cell(row=31, column=13).number_format = '0.0000'
        ws_perb.cell(row=31, column=13).alignment = align_right
        ws_perb.cell(row=31, column=14, value='=IF(M31>0.8,"Sangat Kuat",IF(M31>0.6,"Kuat",IF(M31>0.4,"Cukup Kuat","Lemah")))').alignment = align_center
        
        # WP comparison row
        ws_perb.cell(row=32, column=11, value="SWARA-MAIRCA vs WP").alignment = align_left
        ws_perb.cell(row=32, column=12, value="=P26").alignment = align_right
        ws_perb.cell(row=32, column=12).number_format = '0'
        ws_perb.cell(row=32, column=13, value="=1-(6*L32)/7980")
        ws_perb.cell(row=32, column=13).number_format = '0.0000'
        ws_perb.cell(row=32, column=13).alignment = align_right
        ws_perb.cell(row=32, column=14, value='=IF(M32>0.8,"Sangat Kuat",IF(M32>0.6,"Kuat",IF(M32>0.4,"Cukup Kuat","Lemah")))').alignment = align_center
        
        # TOPSIS comparison row
        ws_perb.cell(row=33, column=11, value="SWARA-MAIRCA vs TOPSIS").alignment = align_left
        ws_perb.cell(row=33, column=12, value="=R26").alignment = align_right
        ws_perb.cell(row=33, column=12).number_format = '0'
        ws_perb.cell(row=33, column=13, value="=1-(6*L33)/7980")
        ws_perb.cell(row=33, column=13).number_format = '0.0000'
        ws_perb.cell(row=33, column=13).alignment = align_right
        ws_perb.cell(row=33, column=14, value='=IF(M33>0.8,"Sangat Kuat",IF(M33>0.6,"Kuat",IF(M33>0.4,"Cukup Kuat","Lemah")))').alignment = align_center
        
        for r_offset in range(3):
            row_num = 31 + r_offset
            for col_idx in range(11, 15):
                cell = ws_perb.cell(row=row_num, column=col_idx)
                cell.font = font_data
                cell.border = box_border
                    
        # Sensitivity Analysis (A s/d E in row 29-35)
        # Determine highest weight criterion dynamically
        top_criterion = swara_rows[0]
        top_code = top_criterion["code"]
        top_idx = c_codes.index(top_code)
        top_w_base = top_criterion["wj"]
        
        ws_perb.cell(row=29, column=1, value=f"Tabel 4. Analisis Sensitivitas (Variasi Bobot {top_code})").font = font_bold
        
        headers_sens = ["Variasi", "Bobot Baru", "Rank 1 Alt.", "Rank Changes", "% Changes"]
        for col_offset, h_text in enumerate(headers_sens):
            cell = ws_perb.cell(row=30, column=1 + col_offset, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = header_border
            
        variations = [
            ("-20%", -0.2),
            ("-10%", -0.1),
            ("Base (0%)", 0.0),
            ("+10%", 0.1),
            ("+20%", 0.2)
        ]
        
        for scen_idx, (scen_name, var_pct) in enumerate(variations):
            row_num = 31 + scen_idx
            ws_perb.row_dimensions[row_num].height = 22
            
            # 1. Calculate new weights locally for python simulation
            w_top_new = top_w_base * (1.0 + var_pct)
            sum_others_base = sum(ordered_weights[j] for j in range(num_criteria) if j != top_idx)
            
            w_new = []
            for j in range(num_criteria):
                if j == top_idx:
                    w_new.append(w_top_new)
                else:
                    w_new.append(ordered_weights[j] * (1.0 - w_top_new) / sum_others_base)
                    
            # 2. Run local MAIRCA simulation with new weights
            sim_qi = []
            for r_idx in range(num_alternatives):
                # PPij = TPij * normalisation
                # TPij = Wj / m
                # Gap = TPij - PPij = (Wj/m) * (1 - normalisation)
                g_sum = 0.0
                for c_idx in range(num_criteria):
                    tp_val = w_new[c_idx] / num_alternatives
                    g_val = tp_val * (1.0 - norm_data[r_idx][c_idx])
                    g_sum += g_val
                sim_qi.append({"alt": alt_codes[r_idx], "qi": g_sum})
                
            sim_sorted = sorted(sim_qi, key=lambda x: x["qi"])
            sim_ranks = {}
            for rank, item in enumerate(sim_sorted, start=1):
                sim_ranks[item["alt"]] = rank
                
            rank_1_alt = sim_sorted[0]["alt"]
            
            # Count rank changes compared to base
            changes_count = 0
            for alt_code in alt_codes:
                if sim_ranks[alt_code] != base_ranks[alt_code]:
                    changes_count += 1
            pct_changes = (changes_count / num_alternatives) * 100.0
            
            # Write to Excel
            ws_perb.cell(row=row_num, column=1, value=scen_name).alignment = align_center
            ws_perb.cell(row=row_num, column=2, value=w_top_new).number_format = '0.0000'
            ws_perb.cell(row=row_num, column=2).alignment = align_right
            ws_perb.cell(row=row_num, column=3, value=rank_1_alt).alignment = align_center
            ws_perb.cell(row=row_num, column=4, value=changes_count).alignment = align_center
            ws_perb.cell(row=row_num, column=5, value=f"{pct_changes:.1f}%").alignment = align_center
            
            for col_offset in range(5):
                cell = ws_perb.cell(row=row_num, column=1 + col_offset)
                cell.font = font_data
                cell.border = box_border
                cell.fill = PatternFill(fill_type=None) # reset fill to none first
                
            if var_pct == 0.0:
                for col_offset in range(5):
                    ws_perb.cell(row=row_num, column=1 + col_offset).fill = fill_base_variation
                    ws_perb.cell(row=row_num, column=1 + col_offset).font = font_bold
                    
        # Autofit column widths for all sheets
        for sheet in wb.worksheets:
            sheet.views.sheetView[0].showGridLines = True
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    if val.startswith('='):
                        val = " 0.0000 "
                    max_len = max(max_len, len(val))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        # Specific overrides for name columns width
        ws_data.column_dimensions['A'].width = 15
        ws_swara.column_dimensions['C'].width = 25
        ws_perb.column_dimensions['K'].width = 25
        
        wb.save(output_path)
        print(f"Dynamic calculations with styled Excel successfully created at: {output_path}")
    except Exception as e:
        print(f"Error running MCDM Engine: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        run_mcdm_engine("referensi/data_smartphone.xlsx", "output/SWARA_MAIRCA_DETAIL_RUMUS.xlsx")
    else:
        run_mcdm_engine(sys.argv[1], sys.argv[2])
